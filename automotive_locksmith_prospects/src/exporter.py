"""Excel export with formatted sheets."""

import logging
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

from config import (
    SHEET_DUPLICATES_REVIEW,
    SHEET_PROSPECTS_CLEAN,
    SHEET_RAW_ABN_FILTERED,
    SHEET_SUBURBS,
    SHEET_SUMMARY_CATEGORY,
    SHEET_SUMMARY_STATE,
    SHEET_SUMMARY_SUBURB,
)

logger = logging.getLogger(__name__)


def _format_sheet(ws: Worksheet, n_rows: int, n_cols: int) -> None:
    """Freeze header, autofilter, bold headers, auto-width columns."""
    if n_rows > 0 and n_cols > 0:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col_idx in range(1, n_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(20, max(10, 12))


def safe_write_dataframe(
    ws: Worksheet,
    df: pd.DataFrame,
    start_row: int = 1,
) -> None:
    """Write dataframe to worksheet, handling large datasets safely."""
    if df.empty:
        return
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == start_row:
                cell.font = Font(bold=True)
    n_rows, n_cols = len(df) + 1, len(df.columns)
    _format_sheet(ws, n_rows, n_cols)


def export_workbook(
    suburbs_df: pd.DataFrame,
    raw_filtered_df: pd.DataFrame,
    prospects_df: pd.DataFrame,
    duplicates_df: pd.DataFrame,
    output_path: Path,
) -> None:
    """
    Create Excel workbook with all required sheets.
    """
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        suburbs_df.to_excel(writer, sheet_name=SHEET_SUBURBS, index=False)
        raw_filtered_df.to_excel(writer, sheet_name=SHEET_RAW_ABN_FILTERED, index=False)
        prospects_df.to_excel(writer, sheet_name=SHEET_PROSPECTS_CLEAN, index=False)
        duplicates_df.to_excel(writer, sheet_name=SHEET_DUPLICATES_REVIEW, index=False)

        # Summary by state
        if not prospects_df.empty and "state" in prospects_df.columns:
            by_state = prospects_df.groupby("state", dropna=False).agg(
                count=("business_name", "count")
            ).reset_index()
            by_state = by_state.sort_values("count", ascending=False)
            by_state.to_excel(writer, sheet_name=SHEET_SUMMARY_STATE, index=False)
        else:
            pd.DataFrame(columns=["state", "count"]).to_excel(
                writer, sheet_name=SHEET_SUMMARY_STATE, index=False
            )

        # Summary by suburb
        if not prospects_df.empty and "suburb" in prospects_df.columns:
            by_suburb = prospects_df.groupby(["suburb", "state"], dropna=False).agg(
                count=("business_name", "count")
            ).reset_index()
            by_suburb = by_suburb.sort_values("count", ascending=False).head(500)
            by_suburb.to_excel(writer, sheet_name=SHEET_SUMMARY_SUBURB, index=False)
        else:
            pd.DataFrame(columns=["suburb", "state", "count"]).to_excel(
                writer, sheet_name=SHEET_SUMMARY_SUBURB, index=False
            )

        # Summary by category
        if not prospects_df.empty and "category" in prospects_df.columns:
            by_cat = prospects_df.groupby("category", dropna=False).agg(
                count=("business_name", "count")
            ).reset_index()
            by_cat = by_cat.sort_values("count", ascending=False)
            by_cat.to_excel(writer, sheet_name=SHEET_SUMMARY_CATEGORY, index=False)
        else:
            pd.DataFrame(columns=["category", "count"]).to_excel(
                writer, sheet_name=SHEET_SUMMARY_CATEGORY, index=False
            )

    # Apply formatting
    wb = load_workbook(path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row > 0 and ws.max_column > 0:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for col_idx in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = min(25, max(10, 15))
    wb.save(path)
    logger.info("Exported workbook to %s", path)
