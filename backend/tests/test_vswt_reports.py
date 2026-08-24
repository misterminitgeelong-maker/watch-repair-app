"""VSWT Regional Intelligence: parser, ranking helpers, and the upload/read API."""
from __future__ import annotations

import io
from uuid import UUID

import openpyxl
import pytest
from sqlmodel import Session

from app.database import engine
from app.models import Tenant, VswtWeeklyShopMetric
from app.routes.vswt_reports import _average, _parse_vswt_workbook, _peer_rows, _rank_of
from app.vswt_kpis import COLUMN_MAP


# ── Workbook builder (mirrors the real VSWT-WSS Summary sheet layout) ──────────────────────

def _build_workbook(week_number, shop_rows: list[dict]) -> bytes:
    """Build an in-memory .xlsx matching the brief's row/column layout for the Summary sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A3"] = "Week Number:"
    ws["C3"] = week_number
    for i, row in enumerate(shop_rows):
        excel_row = 6 + i
        for field, col in COLUMN_MAP:
            if field in row:
                ws.cell(row=excel_row, column=col, value=row[field])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _shop_row(shop_number, shop_name, sales_ty, **extra) -> dict:
    row = {
        "shop_number": shop_number,
        "shop_name": shop_name,
        "area_name": extra.pop("area_name", "VIC SOUTH"),
        "store_format": extra.pop("store_format", "FR"),
        "comp_status": extra.pop("comp_status", "Comp"),
        "sales_ty": sales_ty,
    }
    row.update(extra)
    return row


# ── Parser ───────────────────────────────────────────────────────────────────────────────

def test_parse_detects_week_number_and_shop_rows():
    raw = _build_workbook(32, [
        _shop_row(3269, "Chadstone", 50000),
        _shop_row(3904, "Doncaster", 42000),
    ])
    result = _parse_vswt_workbook(raw, "VSWT-WSS__new_version__32.xlsx")
    assert result["internal_week"] == 32
    assert result["shop_count"] == 2
    numbers = {r["shop_number"] for r in result["rows"]}
    assert numbers == {"3269", "3904"}


def test_parse_skips_rows_with_no_shop_number():
    raw = _build_workbook(32, [
        _shop_row(3269, "Chadstone", 50000),
        {"sales_ty": 99999},  # no shop_number -> not a real shop row
    ])
    result = _parse_vswt_workbook(raw, "f.xlsx")
    assert result["shop_count"] == 1


def test_parse_treats_excel_errors_and_blanks_as_null():
    raw = _build_workbook(32, [
        _shop_row(3269, "Chadstone", "#DIV/0!", jobs_ty="#REF!", customer_ty=""),
    ])
    result = _parse_vswt_workbook(raw, "f.xlsx")
    row = result["rows"][0]
    assert row["sales_ty"] is None
    assert row["jobs_ty"] is None
    assert row["customer_ty"] is None


def test_parse_normalises_shop_number_to_string_without_decimal():
    raw = _build_workbook(32, [_shop_row(3269, "Chadstone", 50000)])
    result = _parse_vswt_workbook(raw, "f.xlsx")
    # openpyxl round-trips a bare int fine, but the source column is genuinely numeric in
    # real exports (float), so shop_number must come out as a clean "3269", not "3269.0".
    assert result["rows"][0]["shop_number"] == "3269"


def test_parse_falls_back_to_first_sheet_when_no_summary_tab():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["C3"] = 5
    ws.cell(row=6, column=2, value=3269)
    ws.cell(row=6, column=9, value=1000)
    buf = io.BytesIO()
    wb.save(buf)
    result = _parse_vswt_workbook(buf.getvalue(), "f.xlsx")
    assert result["shop_count"] == 1


# ── Ranking helpers ──────────────────────────────────────────────────────────────────────

def _metric(shop_number, sales_ty, store_format="FR", comp_status="Comp") -> VswtWeeklyShopMetric:
    return VswtWeeklyShopMetric(
        week_seq=1, shop_number=shop_number, sales_ty=sales_ty,
        store_format=store_format, comp_status=comp_status,
    )


def test_rank_of_higher_is_better_and_ties_share_no_special_casing():
    rows = [_metric("1", 100), _metric("2", 90), _metric("3", 90), _metric("4", 50)]
    assert _rank_of(rows, "sales_ty", 100) == 1
    # two shops ahead at 100 and 90(tie)... shop with 90 has exactly one shop strictly above it (100)
    assert _rank_of(rows, "sales_ty", 90) == 2
    assert _rank_of(rows, "sales_ty", 50) == 4


def test_rank_of_none_value_is_unranked():
    rows = [_metric("1", 100), _metric("2", None)]
    assert _rank_of(rows, "sales_ty", None) is None


def test_average_ignores_nulls():
    assert _average([10, None, 20, None]) == 15
    assert _average([None, None]) is None


def test_peer_rows_filters_franchise_comparable_only():
    rows = [
        _metric("1", 100, store_format="FR", comp_status="Comp"),
        _metric("2", 90, store_format="CO", comp_status="Comp"),
        _metric("3", 80, store_format="FR", comp_status="Non Comp"),
        _metric("4", 70, store_format="FR", comp_status="Comp"),
    ]
    peers = _peer_rows(rows)
    assert {r.shop_number for r in peers} == {"1", "4"}


# ── API: upload -> commit -> read ───────────────────────────────────────────────────────

def _bootstrap(client, tenant_slug, email, role="owner") -> tuple[dict, str]:
    """Bootstrap a tenant (owner) and return (headers, tenant_id)."""
    res = client.post(
        "/v1/auth/bootstrap",
        json={
            "tenant_name": f"Tenant {tenant_slug}",
            "tenant_slug": tenant_slug,
            "owner_email": email,
            "owner_full_name": "Owner",
            "owner_password": "pass123456",
        },
    )
    assert res.status_code == 200, res.text
    tenant_id = res.json()["tenant_id"]
    login = client.post(
        "/v1/auth/login",
        json={"tenant_slug": tenant_slug, "email": email, "password": "pass123456"},
    )
    assert login.status_code == 200, login.text
    return {"Authorization": f"Bearer {login.json()['access_token']}"}, tenant_id


def _set_shop_number(tenant_id: str, shop_number: str) -> None:
    with Session(engine) as session:
        tenant = session.get(Tenant, UUID(tenant_id))
        assert tenant is not None
        tenant.shop_number = shop_number
        session.add(tenant)
        session.commit()


def _make_lower_role_user(client, headers, suffix: str, role: str) -> dict:
    res = client.post(
        "/v1/users",
        headers=headers,
        json={
            "full_name": f"{role.title()} {suffix}",
            "email": f"{role}-{suffix}@test.com",
            "password": "pass123456",
            "role": role,
        },
    )
    assert res.status_code == 201, res.text
    return res.json()


@pytest.fixture
def vswt_client(client):
    return client


def test_upload_then_commit_then_read_flow(vswt_client):
    headers, tenant_id = _bootstrap(vswt_client, "vswt-shop-3269", "owner-3269@test.com")
    _set_shop_number(tenant_id, "3269")

    raw = _build_workbook(41, [
        _shop_row(3269, "Chadstone", 50000, jobs_ty=120, customer_ty=90, budget_sales_target=45000),
        _shop_row(3904, "Doncaster", 70000, jobs_ty=150, customer_ty=110),
        _shop_row(4100, "Bondi", 30000, jobs_ty=60, customer_ty=40, store_format="CO"),
    ])
    upload = vswt_client.post(
        "/v1/reports/vswt/upload",
        headers=headers,
        files=[("files", ("VSWT-WSS__new_version__41.xlsx", raw,
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))],
    )
    assert upload.status_code == 200, upload.text
    batch = upload.json()["batch"]
    assert len(batch) == 1
    assert batch[0]["week_number"] == 41
    assert batch[0]["shop_count"] == 3
    assert batch[0]["overwrite"] is False

    commit = vswt_client.post(
        "/v1/reports/vswt/commit",
        headers=headers,
        json={"batch": [{"filename": batch[0]["filename"], "week_number": 41, "rows": batch[0]["rows"]}]},
    )
    assert commit.status_code == 200, commit.text
    assert commit.json()["saved"] == [{"week_number": 41, "shop_count": 3}]

    summary = vswt_client.get("/v1/reports/vswt/summary", headers=headers)
    assert summary.status_code == 200, summary.text
    body = summary.json()
    assert body["available"] is True
    assert body["shop_number"] == "3269"
    assert body["sales"]["value"] == 50000
    # Doncaster (70000) is the only shop ahead of Chadstone on sales.
    assert body["sales"]["region_rank"] == 2

    rankings = vswt_client.get("/v1/reports/vswt/rankings", headers=headers, params={"week": 41})
    assert rankings.status_code == 200
    sales_row = next(r for r in rankings.json()["rows"] if r["key"] == "sales_ty")
    assert sales_row["region_rank"] == 2
    assert sales_row["region_avg"] == pytest.approx(50000.0)

    scorecard = vswt_client.get("/v1/reports/vswt/scorecard", headers=headers)
    assert scorecard.status_code == 200
    assert scorecard.json()["matrix"][0]["week"] == 41

    leaderboards = vswt_client.get("/v1/reports/vswt/leaderboards", headers=headers, params={"week": 41})
    assert leaderboards.status_code == 200
    sales_board = next(b for b in leaderboards.json()["boards"] if b["key"] == "sales_ty")
    assert sales_board["top"][0]["shop_number"] == "3904"

    trends = vswt_client.get("/v1/reports/vswt/trends", headers=headers)
    assert trends.status_code == 200
    assert trends.json()["sales_series"][-1]["shop"] == 50000


def test_leaderboards_names_every_shop_top_and_bottom(vswt_client):
    """Both top-5 and bottom-5 carry every shop's name/number — this is a private franchisee
    tool, not a public leaderboard, so there's no anonymization to hide who's struggling."""
    headers, tenant_id = _bootstrap(vswt_client, "vswt-shop-bottom", "owner-bottom@test.com")
    _set_shop_number(tenant_id, "2000")

    raw = _build_workbook(41, [
        _shop_row(1000, "Alpha", 90000),
        _shop_row(2000, "Bravo", 10000),  # me — lowest, so I land in "bottom"
        _shop_row(3000, "Charlie", 80000),
        _shop_row(4000, "Delta", 70000),
        _shop_row(5000, "Echo", 60000),
        _shop_row(6000, "Foxtrot", 50000),
    ])
    upload = vswt_client.post(
        "/v1/reports/vswt/upload",
        headers=headers,
        files=[("files", ("VSWT-WSS__new_version__41.xlsx", raw,
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))],
    )
    batch = upload.json()["batch"]
    vswt_client.post(
        "/v1/reports/vswt/commit",
        headers=headers,
        json={"batch": [{"filename": batch[0]["filename"], "week_number": 41, "rows": batch[0]["rows"]}]},
    )

    leaderboards = vswt_client.get("/v1/reports/vswt/leaderboards", headers=headers, params={"week": 41})
    assert leaderboards.status_code == 200
    sales_board = next(b for b in leaderboards.json()["boards"] if b["key"] == "sales_ty")

    # Top-5 stays named.
    assert all(e["shop_number"] is not None and e["shop_name"] is not None for e in sales_board["top"])

    # Bottom-5 is named too now — including my own row (Bravo, is_me).
    assert all(e["shop_number"] is not None and e["shop_name"] is not None for e in sales_board["bottom"])
    me = next(e for e in sales_board["bottom"] if e["is_me"])
    assert me["shop_number"] == "2000"
    assert me["shop_name"] == "Bravo"


def test_reupload_same_week_auto_bumps_to_a_free_week(vswt_client):
    """Auto-assignment always steers clear of a collision (matches the reference app): the
    second upload of an already-stored week number lands on the next free week, it does not
    silently overwrite. Overwriting only happens if the *user* edits the week number back."""
    headers, tenant_id = _bootstrap(vswt_client, "vswt-shop-overwrite", "owner-ow@test.com")
    _set_shop_number(tenant_id, "9001")

    raw1 = _build_workbook(50, [_shop_row(9001, "Shop A", 10000)])
    up1 = vswt_client.post("/v1/reports/vswt/upload", headers=headers,
                            files=[("files", ("w50a.xlsx", raw1))])
    b1 = up1.json()["batch"][0]
    assert b1["week_number"] == 50
    vswt_client.post("/v1/reports/vswt/commit", headers=headers,
                      json={"batch": [{"filename": b1["filename"], "week_number": 50, "rows": b1["rows"]}]})

    raw2 = _build_workbook(50, [_shop_row(9001, "Shop A", 20000)])
    up2 = vswt_client.post("/v1/reports/vswt/upload", headers=headers,
                            files=[("files", ("w50b.xlsx", raw2))])
    b2 = up2.json()["batch"][0]
    assert b2["internal_week"] == 50
    assert b2["week_number"] == 51  # bumped, not overwritten
    assert b2["overwrite"] is False
    assert 50 in up2.json()["existing_weeks"]  # frontend uses this to warn on a manual edit


def test_commit_with_an_existing_week_number_overwrites_not_duplicates(vswt_client):
    """The commit step itself (delete-then-insert per week) is what actually overwrites —
    exercised here as if the user had edited the confirm step's week number back onto week 50."""
    headers, tenant_id = _bootstrap(vswt_client, "vswt-shop-manual-overwrite", "owner-mo@test.com")
    _set_shop_number(tenant_id, "9002")

    raw1 = _build_workbook(50, [_shop_row(9002, "Shop A", 10000)])
    up1 = vswt_client.post("/v1/reports/vswt/upload", headers=headers,
                            files=[("files", ("w50a.xlsx", raw1))])
    b1 = up1.json()["batch"][0]
    vswt_client.post("/v1/reports/vswt/commit", headers=headers,
                      json={"batch": [{"filename": b1["filename"], "week_number": 50, "rows": b1["rows"]}]})

    raw2 = _build_workbook(50, [_shop_row(9002, "Shop A", 20000)])
    up2 = vswt_client.post("/v1/reports/vswt/upload", headers=headers,
                            files=[("files", ("w50b.xlsx", raw2))])
    b2 = up2.json()["batch"][0]
    # User corrects the auto-bumped number back to 50 in the confirm step.
    vswt_client.post("/v1/reports/vswt/commit", headers=headers,
                      json={"batch": [{"filename": b2["filename"], "week_number": 50, "rows": b2["rows"]}]})

    summary = vswt_client.get("/v1/reports/vswt/summary", headers=headers)
    assert summary.json()["sales"]["value"] == 20000  # overwritten, not duplicated

    weeks = vswt_client.get("/v1/reports/vswt/weeks", headers=headers)
    week_50 = next(w for w in weeks.json()["weeks"] if w["week"] == 50)
    assert week_50["shop_count"] == 1


def test_commit_rejects_duplicate_week_numbers_in_one_batch(vswt_client):
    headers, _tid = _bootstrap(vswt_client, "vswt-dup-batch", "owner-dup@test.com")
    raw = _build_workbook(9, [_shop_row(1, "A", 100)])
    res = vswt_client.post(
        "/v1/reports/vswt/commit",
        headers=headers,
        json={"batch": [
            {"filename": "a.xlsx", "week_number": 9, "rows": [_shop_row(1, "A", 100)]},
            {"filename": "b.xlsx", "week_number": 9, "rows": [_shop_row(2, "B", 200)]},
        ]},
    )
    assert res.status_code == 400


def test_upload_requires_manager_or_above(vswt_client):
    headers, _tid = _bootstrap(vswt_client, "vswt-role-gate", "owner-rg@test.com")
    intake = _make_lower_role_user(vswt_client, headers, "rg", "intake")
    login = vswt_client.post(
        "/v1/auth/login",
        json={"tenant_slug": "vswt-role-gate", "email": intake["email"], "password": "pass123456"},
    )
    intake_headers = {"Authorization": f"Bearer {login.json()['access_token']}"}

    raw = _build_workbook(12, [_shop_row(1, "A", 100)])
    res = vswt_client.post("/v1/reports/vswt/upload", headers=intake_headers,
                            files=[("files", ("f.xlsx", raw))])
    assert res.status_code == 403


def test_summary_unavailable_without_shop_number(vswt_client):
    headers, _tid = _bootstrap(vswt_client, "vswt-no-shop-number", "owner-ns@test.com")
    res = vswt_client.get("/v1/reports/vswt/summary", headers=headers)
    assert res.status_code == 200
    body = res.json()
    assert body["available"] is False
    assert body["reason"] == "no_shop_number"


def test_summary_unavailable_when_shop_not_in_data(vswt_client):
    headers, tenant_id = _bootstrap(vswt_client, "vswt-shop-missing", "owner-sm@test.com")
    _set_shop_number(tenant_id, "5555")

    raw = _build_workbook(60, [_shop_row(1, "A", 100)])
    up = vswt_client.post("/v1/reports/vswt/upload", headers=headers,
                           files=[("files", ("f.xlsx", raw))])
    b = up.json()["batch"][0]
    vswt_client.post("/v1/reports/vswt/commit", headers=headers,
                      json={"batch": [{"filename": b["filename"], "week_number": 60, "rows": b["rows"]}]})

    res = vswt_client.get("/v1/reports/vswt/summary", headers=headers)
    body = res.json()
    assert body["available"] is False
    assert body["reason"] == "shop_not_found"


# ── Directory + browsing other shops ────────────────────────────────────────────────────

def _seed_directory_week(vswt_client, headers, week: int) -> None:
    raw = _build_workbook(week, [
        _shop_row(3269, "Chadstone", 62000, area_name="VIC SOUTH", store_format="FR", comp_status="Comp"),
        _shop_row(3904, "Doncaster", 81000, area_name="VIC SOUTH", store_format="FR", comp_status="Comp"),
        _shop_row(4100, "Bondi", 45000, area_name="NSW EAST", store_format="FR", comp_status="Comp"),
        _shop_row(4200, "Chatswood", 71000, area_name="NSW EAST", store_format="CO", comp_status="Comp"),
    ])
    up = vswt_client.post("/v1/reports/vswt/upload", headers=headers, files=[("files", ("d.xlsx", raw))])
    b = up.json()["batch"][0]
    commit = vswt_client.post(
        "/v1/reports/vswt/commit", headers=headers,
        json={"batch": [{"filename": b["filename"], "week_number": week, "rows": b["rows"]}]},
    )
    assert commit.status_code == 200, commit.text


def _seed_weeks(vswt_client, headers, week_numbers, rows_fn) -> None:
    """Upload+commit one week per number in `week_numbers`, rows for each built by `rows_fn(week)
    -> list[shop_row dict]` — for tests that need several weeks of history (Shop Report windows,
    consistency leaderboards)."""
    for w in week_numbers:
        raw = _build_workbook(w, rows_fn(w))
        up = vswt_client.post("/v1/reports/vswt/upload", headers=headers, files=[("files", (f"w{w}.xlsx", raw))])
        b = up.json()["batch"][0]
        commit = vswt_client.post(
            "/v1/reports/vswt/commit", headers=headers,
            json={"batch": [{"filename": b["filename"], "week_number": w, "rows": b["rows"]}]},
        )
        assert commit.status_code == 200, commit.text


def test_directory_lists_every_shop_and_flags_me_and_peers(vswt_client):
    headers, tenant_id = _bootstrap(vswt_client, "vswt-dir-basic", "owner-dirbasic@test.com")
    _set_shop_number(tenant_id, "3269")
    _seed_directory_week(vswt_client, headers, 70)

    res = vswt_client.get("/v1/reports/vswt/directory", headers=headers, params={"week": 70})
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["available"] is True
    assert body["region_size"] == 4
    assert {r["shop_number"] for r in body["rows"]} == {"3269", "3904", "4100", "4200"}
    me = next(r for r in body["rows"] if r["shop_number"] == "3269")
    assert me["is_me"] is True
    assert me["is_peer"] is True
    chatswood = next(r for r in body["rows"] if r["shop_number"] == "4200")
    assert chatswood["is_me"] is False
    assert chatswood["is_peer"] is False  # CO, not FR+Comp
    assert "sales_ty" in me["values"]  # Headline group by default


def test_directory_search_filters_by_name_number_or_area(vswt_client):
    headers, tenant_id = _bootstrap(vswt_client, "vswt-dir-search", "owner-dirsearch@test.com")
    _set_shop_number(tenant_id, "3269")
    _seed_directory_week(vswt_client, headers, 71)

    by_name = vswt_client.get("/v1/reports/vswt/directory", headers=headers, params={"week": 71, "search": "donc"})
    assert {r["shop_number"] for r in by_name.json()["rows"]} == {"3904"}

    by_number = vswt_client.get("/v1/reports/vswt/directory", headers=headers, params={"week": 71, "search": "4100"})
    assert {r["shop_number"] for r in by_number.json()["rows"]} == {"4100"}

    by_area = vswt_client.get("/v1/reports/vswt/directory", headers=headers, params={"week": 71, "search": "nsw east"})
    assert {r["shop_number"] for r in by_area.json()["rows"]} == {"4100", "4200"}


def test_directory_peer_only_excludes_non_franchise_comp(vswt_client):
    headers, tenant_id = _bootstrap(vswt_client, "vswt-dir-peer", "owner-dirpeer@test.com")
    _set_shop_number(tenant_id, "3269")
    _seed_directory_week(vswt_client, headers, 72)

    res = vswt_client.get(
        "/v1/reports/vswt/directory", headers=headers, params={"week": 72, "peer_only": True}
    )
    assert {r["shop_number"] for r in res.json()["rows"]} == {"3269", "3904", "4100"}  # excludes Chatswood (CO)


def test_directory_requires_own_shop_number(vswt_client):
    headers, _tid = _bootstrap(vswt_client, "vswt-dir-noshop", "owner-dirnoshop@test.com")
    res = vswt_client.get("/v1/reports/vswt/directory", headers=headers)
    body = res.json()
    assert body["available"] is False
    assert body["reason"] == "no_shop_number"


def test_can_browse_another_shops_rankings_scorecard_and_trends(vswt_client):
    headers, tenant_id = _bootstrap(vswt_client, "vswt-browse", "owner-browse@test.com")
    _set_shop_number(tenant_id, "3269")
    _seed_directory_week(vswt_client, headers, 73)

    rankings = vswt_client.get(
        "/v1/reports/vswt/rankings", headers=headers, params={"week": 73, "shop_number": "3904"}
    )
    assert rankings.status_code == 200
    rb = rankings.json()
    assert rb["available"] is True
    assert rb["shop_number"] == "3904"
    assert rb["shop_name"] == "Doncaster"
    assert rb["viewing_own_shop"] is False
    sales_row = next(r for r in rb["rows"] if r["key"] == "sales_ty")
    assert sales_row["value"] == 81000

    scorecard = vswt_client.get(
        "/v1/reports/vswt/scorecard", headers=headers, params={"shop_number": "3904"}
    )
    sb = scorecard.json()
    assert sb["available"] is True
    assert sb["shop_number"] == "3904"
    assert sb["viewing_own_shop"] is False

    trends = vswt_client.get(
        "/v1/reports/vswt/trends", headers=headers, params={"shop_number": "3904"}
    )
    tb = trends.json()
    assert tb["available"] is True
    assert tb["shop_number"] == "3904"
    assert tb["viewing_own_shop"] is False
    assert tb["sales_series"][-1]["shop"] == 81000

    # Omitting shop_number still defaults to your own shop.
    own = vswt_client.get("/v1/reports/vswt/rankings", headers=headers, params={"week": 73})
    assert own.json()["shop_number"] == "3269"
    assert own.json()["viewing_own_shop"] is True


# ── Shop Report (Week / Month / Year windows) ───────────────────────────────────────────
#
# These tests all use `watch_sales_ty` (Category Sales group) rather than the usual `sales_ty` —
# no other test in this file ever sets that field, so it's a clean, unpolluted metric to compute
# region-wide averages/ranks over in this shared (not per-test-isolated) table. They also use
# shop numbers and week numbers not used anywhere else in this file, for the same reason: a
# shop/week only ever shows up in accumulations this test itself seeded.

def test_shop_report_week_month_year_windows_and_consistency_stats(vswt_client):
    """Consistent Co posts a steady $20k of watch sales every week for 10 weeks. Big Week Co
    posts $5k every week except one $100k blowout in the final (latest) week. That one week is
    enough to put Big Week Co ahead on the Week view, and even still ahead on the rolling Month
    average (only 4 weeks, so the blowout dominates) — but over the full Year average, Consistent
    Co's steadiness wins."""
    headers, tenant_id = _bootstrap(vswt_client, "vswt-shop-report", "owner-sr@test.com")
    _set_shop_number(tenant_id, "800001")

    def rows(w):
        rival_sales = 100000 if w == 900010 else 5000
        return [
            _shop_row(800001, "Consistent Co", 0, watch_sales_ty=20000),
            _shop_row(800002, "Big Week Co", 0, watch_sales_ty=rival_sales),
        ]

    _seed_weeks(vswt_client, headers, range(900001, 900011), rows)

    res = vswt_client.get(
        "/v1/reports/vswt/shop-report", headers=headers, params={"group": "Category Sales"}
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["available"] is True
    assert body["shop_number"] == "800001"
    assert body["viewing_own_shop"] is True
    assert body["weeks_tracked"] >= 10  # this table is shared across the whole test session
    assert body["region_size"] == 2

    row = next(r for r in body["rows"] if r["key"] == "watch_sales_ty")
    assert row["week"] == {"value": 20000, "rank": 2}
    assert row["month"]["value"] == pytest.approx(20000.0)
    assert row["month"]["rank"] == 2
    assert row["month"]["weeks_counted"] == 4
    assert row["year"]["value"] == pytest.approx(20000.0)
    assert row["year"]["rank"] == 1  # consistency wins over one huge week
    assert row["year"]["weeks_counted"] == 10
    assert row["year"]["best_rank"] == 1
    assert row["year"]["worst_rank"] == 2
    assert row["year"]["rank_stdev"] == pytest.approx(0.3)

    rival = vswt_client.get(
        "/v1/reports/vswt/shop-report", headers=headers,
        params={"shop_number": "800002", "group": "Category Sales"},
    )
    rb = rival.json()
    assert rb["available"] is True
    assert rb["viewing_own_shop"] is False
    rival_row = next(r for r in rb["rows"] if r["key"] == "watch_sales_ty")
    assert rival_row["week"] == {"value": 100000, "rank": 1}
    assert rival_row["year"]["value"] == pytest.approx(14500.0)
    assert rival_row["year"]["rank"] == 2
    assert rival_row["year"]["best_rank"] == 1
    assert rival_row["year"]["worst_rank"] == 2


def test_shop_report_unavailable_without_shop_number(vswt_client):
    headers, _tid = _bootstrap(vswt_client, "vswt-sr-no-shop", "owner-srns@test.com")
    res = vswt_client.get("/v1/reports/vswt/shop-report", headers=headers)
    body = res.json()
    assert body["available"] is False
    assert body["reason"] == "no_shop_number"


def test_shop_report_unavailable_when_shop_not_in_data(vswt_client):
    headers, tenant_id = _bootstrap(vswt_client, "vswt-sr-missing", "owner-srm@test.com")
    _set_shop_number(tenant_id, "800099")  # never appears in any seeded week
    _seed_directory_week(vswt_client, headers, 900020)

    res = vswt_client.get("/v1/reports/vswt/shop-report", headers=headers)
    body = res.json()
    assert body["available"] is False
    assert body["reason"] == "shop_not_found"


# ── Leaderboards: Consistency mode ──────────────────────────────────────────────────────

def test_leaderboards_consistency_mode_ranks_by_average_rank_not_one_week(vswt_client):
    """A shop with one huge week should not top the Consistency board over a shop that's
    reliably ahead every other week — that's the whole point of the mode. Uses `key_sales_ty`
    (distinct from the `watch_sales_ty` field used above) so this test's own shops don't mix into
    the same ranking pool as the Shop Report test's, since Consistency boards aggregate across
    every week on file regardless of which test wrote it."""
    headers, tenant_id = _bootstrap(vswt_client, "vswt-lb-consistency", "owner-lbc@test.com")
    _set_shop_number(tenant_id, "800011")

    def rows(w):
        rival_sales = 100000 if w == 900040 else 5000
        return [
            _shop_row(800011, "Consistent Co", 0, key_sales_ty=20000),
            _shop_row(800012, "Big Week Co", 0, key_sales_ty=rival_sales),
        ]

    _seed_weeks(vswt_client, headers, range(900031, 900041), rows)

    res = vswt_client.get(
        "/v1/reports/vswt/leaderboards", headers=headers,
        params={"mode": "consistency", "group": "Category Sales"},
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["available"] is True
    assert body["mode"] == "consistency"

    board = next(b for b in body["boards"] if b["key"] == "key_sales_ty")
    assert board["type"] == "ratio"
    assert [e["shop_number"] for e in board["top"]] == ["800011", "800012"]
    assert board["top"][0]["value"] == pytest.approx(1.1)  # rank 1 for 9 weeks, rank 2 once
    assert board["top"][0]["weeks_counted"] == 10
    assert board["top"][1]["value"] == pytest.approx(1.9)
    assert board["bottom"] == []  # only 2 qualifying shops, both already shown in "top"
    assert board["my_rank"] == 1


def test_leaderboards_consistency_requires_minimum_weeks_to_qualify(vswt_client):
    """A shop that's only ever appeared once shouldn't be able to claim a top consistency spot
    off a single lucky week. Uses `engrave_sales_ty` so it doesn't share a ranking pool with the
    other Consistency-mode test above."""
    headers, tenant_id = _bootstrap(vswt_client, "vswt-lb-minweeks", "owner-lbmw@test.com")
    _set_shop_number(tenant_id, "800021")

    def rows(w):
        base = [
            _shop_row(800021, "Consistent Co", 0, engrave_sales_ty=20000),
            _shop_row(800022, "Regular", 0, engrave_sales_ty=15000),
        ]
        if w == 900055:  # only ever uploaded once
            base.append(_shop_row(800023, "One Week Wonder", 0, engrave_sales_ty=999999))
        return base

    _seed_weeks(vswt_client, headers, range(900051, 900056), rows)

    res = vswt_client.get(
        "/v1/reports/vswt/leaderboards", headers=headers,
        params={"mode": "consistency", "group": "Category Sales"},
    )
    board = next(b for b in res.json()["boards"] if b["key"] == "engrave_sales_ty")
    shop_numbers = {e["shop_number"] for e in board["top"] + board["bottom"]}
    assert "800023" not in shop_numbers
    assert board["total"] == 2  # only the two shops with >= min(3, len(weeks)) weeks qualify


def test_leaderboards_latest_mode_is_still_the_default(vswt_client):
    headers, tenant_id = _bootstrap(vswt_client, "vswt-lb-default-mode", "owner-lbdm@test.com")
    _set_shop_number(tenant_id, "3269")
    _seed_directory_week(vswt_client, headers, 90)

    res = vswt_client.get("/v1/reports/vswt/leaderboards", headers=headers, params={"week": 90})
    assert res.json()["mode"] == "latest"


# ── Weekly report builder ───────────────────────────────────────────────────────────────

def test_weekly_report_returns_picked_shops_sorted_by_sales_with_totals(vswt_client):
    headers, tenant_id = _bootstrap(vswt_client, "vswt-wr-basic", "owner-wrbasic@test.com")
    _set_shop_number(tenant_id, "3269")
    _seed_directory_week(vswt_client, headers, 200)  # Chadstone 62k, Doncaster 81k, Bondi 45k, Chatswood 71k

    res = vswt_client.get(
        "/v1/reports/vswt/weekly-report", headers=headers,
        params={"week": 200, "shop_numbers": "4100,3269,3904"},  # deliberately out of sales order
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["available"] is True
    assert body["week"] == 200
    assert body["region_size"] == 4
    # Comprehensive: every KPI group is present, not just Headline.
    assert set(body["groups"]) >= {"Headline", "Conversion", "Category Sales"}
    assert len(body["kpis"]) > 6  # every KPI across every group, not one group's worth

    # Sorted best sales first, regardless of the order shop_numbers were passed in.
    assert [s["shop_number"] for s in body["shops"]] == ["3904", "3269", "4100"]
    assert body["shops"][0]["sales_value"] == 81000
    assert body["shops"][0]["sales_rank"] == 1  # Doncaster: highest sales in the whole 4-shop region
    me = next(s for s in body["shops"] if s["shop_number"] == "3269")
    assert me["is_me"] is True

    # Every KPI carries both a value and a region rank, plus a composite overall_avg_rank.
    doncaster = body["shops"][0]
    assert doncaster["ranks"]["sales_ty"] == 1
    assert doncaster["values"]["sales_ty"] == 81000
    assert doncaster["overall_avg_rank"] is not None

    assert body["totals"]["sales"] == 62000 + 81000 + 45000
    assert body["missing_shop_numbers"] == []


def test_weekly_report_reports_missing_shop_numbers_without_failing(vswt_client):
    headers, tenant_id = _bootstrap(vswt_client, "vswt-wr-missing", "owner-wrmissing@test.com")
    _set_shop_number(tenant_id, "3269")
    _seed_directory_week(vswt_client, headers, 201)

    res = vswt_client.get(
        "/v1/reports/vswt/weekly-report", headers=headers,
        params={"week": 201, "shop_numbers": "3269,9999999"},
    )
    body = res.json()
    assert body["available"] is True
    assert [s["shop_number"] for s in body["shops"]] == ["3269"]
    assert body["missing_shop_numbers"] == ["9999999"]


def test_weekly_report_requires_at_least_one_shop(vswt_client):
    headers, tenant_id = _bootstrap(vswt_client, "vswt-wr-empty", "owner-wrempty@test.com")
    _set_shop_number(tenant_id, "3269")
    _seed_directory_week(vswt_client, headers, 202)

    res = vswt_client.get(
        "/v1/reports/vswt/weekly-report", headers=headers,
        params={"week": 202, "shop_numbers": "  , ,"},
    )
    assert res.status_code == 400


def test_weekly_report_pdf_downloads_as_pdf(vswt_client):
    headers, tenant_id = _bootstrap(vswt_client, "vswt-wr-pdf", "owner-wrpdf@test.com")
    _set_shop_number(tenant_id, "3269")
    _seed_directory_week(vswt_client, headers, 203)

    res = vswt_client.get(
        "/v1/reports/vswt/weekly-report/pdf", headers=headers,
        params={"week": 203, "shop_numbers": "3269,3904", "title": "Change Makers Weekly Report"},
    )
    assert res.status_code == 200, res.text
    assert res.headers["content-type"] == "application/pdf"
    assert "attachment" in res.headers["content-disposition"]
    assert res.content[:4] == b"%PDF"


def test_weekly_report_pdf_404s_when_no_picked_shop_is_found(vswt_client):
    headers, tenant_id = _bootstrap(vswt_client, "vswt-wr-pdf-404", "owner-wrpdf404@test.com")
    _set_shop_number(tenant_id, "3269")
    _seed_directory_week(vswt_client, headers, 204)

    res = vswt_client.get(
        "/v1/reports/vswt/weekly-report/pdf", headers=headers,
        params={"week": 204, "shop_numbers": "9999999"},
    )
    assert res.status_code == 404


def test_weekly_report_compare_within_selection_reranks_against_the_picked_shops_only(vswt_client):
    """Region-wide, Chadstone (62k) is #3 of 4 (behind Doncaster 81k and Chatswood 71k). Leave
    Doncaster out of the report and Chadstone should jump to #1 among just the picked shops."""
    headers, tenant_id = _bootstrap(vswt_client, "vswt-wr-within", "owner-wrwithin@test.com")
    _set_shop_number(tenant_id, "3269")
    _seed_directory_week(vswt_client, headers, 205)  # Chadstone 62k, Bondi 45k, Chatswood 71k (Doncaster 81k excluded)

    region_res = vswt_client.get(
        "/v1/reports/vswt/weekly-report", headers=headers,
        params={"week": 205, "shop_numbers": "3269,4100,4200"},
    )
    region_body = region_res.json()
    assert region_body["compare_within_selection"] is False
    assert region_body["rank_pool_size"] == 4  # whole region, including Doncaster who isn't even in the report
    chadstone_region = next(s for s in region_body["shops"] if s["shop_number"] == "3269")
    assert chadstone_region["sales_rank"] == 3  # behind Doncaster (81k) and Chatswood (71k) region-wide

    within_res = vswt_client.get(
        "/v1/reports/vswt/weekly-report", headers=headers,
        params={"week": 205, "shop_numbers": "3269,4100,4200", "compare_within_selection": True},
    )
    within_body = within_res.json()
    assert within_body["compare_within_selection"] is True
    assert within_body["rank_pool_size"] == 3  # just the 3 picked shops
    chadstone_within = next(s for s in within_body["shops"] if s["shop_number"] == "3269")
    assert chadstone_within["sales_rank"] == 2  # still behind Chatswood, but now out of 3 not 4
    bondi_within = next(s for s in within_body["shops"] if s["shop_number"] == "4100")
    assert bondi_within["sales_rank"] == 3  # last of the 3 picked shops (lowest sales)


def test_weekly_report_pdf_compare_within_selection_downloads_fine(vswt_client):
    headers, tenant_id = _bootstrap(vswt_client, "vswt-wr-within-pdf", "owner-wrwithinpdf@test.com")
    _set_shop_number(tenant_id, "3269")
    _seed_directory_week(vswt_client, headers, 206)

    res = vswt_client.get(
        "/v1/reports/vswt/weekly-report/pdf", headers=headers,
        params={"week": 206, "shop_numbers": "3269,4100", "compare_within_selection": True},
    )
    assert res.status_code == 200, res.text
    assert res.content[:4] == b"%PDF"
