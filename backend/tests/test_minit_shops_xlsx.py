"""Unit tests for Minit TSS xlsx parsing."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from app.minit_shops import (
    derive_au_state_from_area_region,
    parse_minit_shops_xlsx,
    parse_minit_shops_xlsx_detailed,
    tenant_slug_for_shop,
)


def _write_sample_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "TSS Scores"
    ws.append(["note row"])
    ws.append(["2025-12-31"])
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append(["Shop #", "Shop Name", "Area", "Region"])
    ws.append(["Raw Score", "TSS score"])  # subheader row — skipped
    ws.append([3269, "Chadstone", "VIC SOUTH", "VIC"])
    ws.append([4278.0, "Toowoomba", "QLD WEST", "QLD"])
    ws.append([6012, "Whitford City", "WA NORTH", "SW"])
    ws.append([9999, "Duplicate", "X", "NSW"])
    ws.append([9999, "Duplicate Again", "X", "NSW"])
    wb.save(path)
    wb.close()


def test_parse_minit_shops_xlsx_skips_subheaders_and_duplicates(tmp_path: Path) -> None:
    xlsx = tmp_path / "sample.xlsx"
    _write_sample_xlsx(xlsx)
    shops = parse_minit_shops_xlsx(xlsx)
    assert len(shops) == 4
    by_num = {s.shop_number: s for s in shops}
    assert by_num["3269"].name == "Chadstone"
    assert by_num["3269"].region == "VIC"
    assert by_num["3269"].area == "VIC SOUTH"
    assert by_num["3269"].business_address == "Chadstone, VIC SOUTH, VIC"
    assert by_num["3269"].state_code == "VIC"
    assert by_num["4278"].name == "Toowoomba"
    assert by_num["6012"].state_code == "WA"
    assert tenant_slug_for_shop(by_num["3269"]) == "minit-3269"


def test_derive_au_state_from_area_region() -> None:
    assert derive_au_state_from_area_region("VIC SOUTH", "VIC") == "VIC"
    assert derive_au_state_from_area_region("WA NORTH", "SW") == "WA"
    assert derive_au_state_from_area_region("SOUTH AUSTRALIA", "SW") == "SA"
    assert derive_au_state_from_area_region("MALAYSIA", "SEA") is None
    assert derive_au_state_from_area_region("NZ NORTH", "NZ") is None


def _write_claude_shops_sheet_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Shops"
    ws.append(["Shop #", "Shop Name", "Area", "Region"])
    ws.append([4278, "Toowoomba", "QLD WEST", "QLD"])
    ws.append([9999, "First", "NSW", "NSW"])
    ws.append([9999, "Duplicate", "NSW", "NSW"])
    ws.append(["bad-id", "Bad Shop", "X", "X"])
    ws.append([8888, "", "QLD", "QLD"])
    wb.save(path)
    wb.close()


def test_parse_claude_shops_sheet_format(tmp_path: Path) -> None:
    xlsx = tmp_path / "claude_shops.xlsx"
    _write_claude_shops_sheet_xlsx(xlsx)
    result = parse_minit_shops_xlsx_detailed(xlsx, collect_row_errors=True)
    assert result.sheet_name == "Shops"
    assert len(result.shops) == 2
    by_num = {s.shop_number: s for s in result.shops}
    assert by_num["4278"].name == "Toowoomba"
    assert by_num["9999"].name == "First"
    assert any("duplicate" in e.lower() for e in result.errors)
    assert any("invalid shop number" in e.lower() for e in result.errors)
    assert any("8888" in e for e in result.errors)


def test_parse_minit_shops_xlsx_missing_header_raises(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["No shop column here"])
    path = tmp_path / "bad.xlsx"
    wb.save(path)
    wb.close()
    with pytest.raises(ValueError, match="Shop #"):
        parse_minit_shops_xlsx(path)
