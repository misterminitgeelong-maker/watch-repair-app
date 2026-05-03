import csv
import io
from datetime import datetime, timezone
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from pydantic import BaseModel
from sqlmodel import Session, select

from ..database import get_session
from ..dependencies import AuthContext, get_auth_context, require_tech_or_above
from ..models import Customer, CustomerOrder, CustomerOrderCreate, CustomerOrderRead, CustomerOrderUpdate

router = APIRouter(
    prefix="/v1/customer-orders",
    tags=["customer-orders"],
)


def _to_read(order: CustomerOrder, session: Session) -> CustomerOrderRead:
    customer_name: str | None = None
    if order.customer_id:
        customer = session.get(Customer, order.customer_id)
        if customer:
            customer_name = customer.full_name
    data = order.model_dump()
    data["customer_name"] = customer_name
    return CustomerOrderRead(**data)


@router.get("", response_model=list[CustomerOrderRead])
def list_customer_orders(
    status: str | None = Query(default=None),
    session: Session = Depends(get_session),
    auth: AuthContext = Depends(get_auth_context),
):
    q = select(CustomerOrder).where(CustomerOrder.tenant_id == auth.tenant_id)
    if status:
        q = q.where(CustomerOrder.status == status)
    q = q.order_by(CustomerOrder.created_at.desc())
    orders = session.exec(q).all()
    return [_to_read(o, session) for o in orders]


@router.post("", response_model=CustomerOrderRead)
def create_customer_order(
    payload: CustomerOrderCreate,
    session: Session = Depends(get_session),
    auth: AuthContext = Depends(get_auth_context),
    _: None = Depends(require_tech_or_above),
):
    order = CustomerOrder(
        tenant_id=auth.tenant_id,
        **payload.model_dump(),
    )
    session.add(order)
    session.commit()
    session.refresh(order)
    return _to_read(order, session)


@router.patch("/{order_id}", response_model=CustomerOrderRead)
def update_customer_order(
    order_id: UUID,
    payload: CustomerOrderUpdate,
    session: Session = Depends(get_session),
    auth: AuthContext = Depends(get_auth_context),
    _: None = Depends(require_tech_or_above),
):
    order = session.get(CustomerOrder, order_id)
    if not order or order.tenant_id != auth.tenant_id:
        raise HTTPException(status_code=404, detail="Order not found")
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(order, field, value)
    order.updated_at = datetime.now(timezone.utc)
    session.add(order)
    session.commit()
    session.refresh(order)
    return _to_read(order, session)


@router.delete("/{order_id}", status_code=204)
def delete_customer_order(
    order_id: UUID,
    session: Session = Depends(get_session),
    auth: AuthContext = Depends(get_auth_context),
    _: None = Depends(require_tech_or_above),
):
    order = session.get(CustomerOrder, order_id)
    if not order or order.tenant_id != auth.tenant_id:
        raise HTTPException(status_code=404, detail="Order not found")
    session.delete(order)
    session.commit()


# ── Import ────────────────────────────────────────────────────────────────────

class CustomerOrderImportResult(BaseModel):
    dry_run: bool
    total_rows: int
    imported: int
    skipped: int
    skipped_reasons: dict[str, int]


_VALID_STATUSES = {"to_order", "ordered", "arrived", "notified", "collected"}
_VALID_PRIORITIES = {"normal", "high", "urgent"}

_GANG_COLLECTED = {"collected", "c", "collect"}
_GANG_NO_GO = {"ng", "no go", "no_go", "cancelled", "canceled"}
_GANG_GO_AHEAD = {"ga", "go ahead", "go_ahead", "approved", "yes", "invoice"}
# "Ordered date" values that mean the item is physically in the shop
_ARRIVED_WORDS = {
    "here", "complete", "complete.", "ready", "arrived", "in stock",
    "in drawer", "in_drawer", "drawer",
}


def _normalise(raw: str) -> str:
    return raw.strip().lower().replace(" ", "_")


def _parse_cents(raw: str) -> int:
    """Parse cost strings like '$35.10 + gst', '24.95', 'nc', 'Paid' → cents."""
    import re
    raw = raw.strip()
    if not raw or raw.lower() in {"nc", "n/c", "paid", "free", "no charge", "nil", "tbc", "quote"}:
        return 0
    raw = re.sub(r"\$", "", raw)
    raw = re.sub(r"\s*\+\s*gst.*", "", raw, flags=re.IGNORECASE)
    raw = raw.replace(",", "").strip()
    try:
        return max(0, round(float(raw) * 100))
    except (ValueError, TypeError):
        return 0


def _derive_status_and_supplier(gang: str, ordered_date: str) -> tuple[str, str | None]:
    """Return (status, supplier_hint) from GA/NG and Ordered date columns.

    When 'Ordered date' holds a supplier name (Leffler, LSC, Remote King…) rather
    than an arrival keyword, we capture it as the supplier and mark status=ordered.
    """
    gang_l = gang.strip().lower()
    ordered_l = ordered_date.strip().lower()
    ordered_raw = ordered_date.strip() or None

    if gang_l in _GANG_COLLECTED or gang_l in _GANG_NO_GO:
        return "collected", None

    if ordered_l in _ARRIVED_WORDS:
        return "arrived", None

    if gang_l in _GANG_GO_AHEAD or gang_l in _VALID_STATUSES:
        if ordered_l:
            # Non-arrival text = supplier name
            return "ordered", ordered_raw
        return "ordered", None

    # Generic CSV status column
    if gang_l in _VALID_STATUSES:
        return gang_l, None

    return "to_order", None


class _CustomerCache:
    """Load all tenant customers once; resolve/create without extra DB round-trips."""

    def __init__(self, session: Session, tenant_id: UUID) -> None:
        self._session = session
        self._tenant_id = tenant_id
        all_customers = session.exec(
            select(Customer).where(Customer.tenant_id == tenant_id)
        ).all()
        # Index by normalised name and normalised phone (spaces stripped)
        self._by_name: dict[str, Customer] = {}
        self._by_phone: dict[str, Customer] = {}
        for c in all_customers:
            if c.full_name:
                self._by_name[c.full_name.strip().lower()] = c
            if c.phone:
                self._by_phone[c.phone.replace(" ", "")] = c

    @staticmethod
    def _clean_phone(raw: str) -> str:
        raw = raw.strip().rstrip("0").rstrip(".") if raw.endswith(".0") else raw.strip()
        return raw.replace(" ", "")

    def find_or_create(self, name: str, phone: str) -> UUID | None:
        name = name.strip()
        phone_clean = self._clean_phone(phone)
        if not name and not phone_clean:
            return None

        # Phone match first (most specific)
        if phone_clean and phone_clean in self._by_phone:
            return self._by_phone[phone_clean].id

        # Name match
        if name and name.lower() in self._by_name:
            c = self._by_name[name.lower()]
            return c.id

        # Create new
        if name:
            new_c = Customer(
                tenant_id=self._tenant_id,
                full_name=name,
                phone=phone.strip() or None,
            )
            self._session.add(new_c)
            self._session.flush()
            self._by_name[name.lower()] = new_c
            if phone_clean:
                self._by_phone[phone_clean] = new_c
            return new_c.id

        return None


def _read_csv_rows(content: bytes) -> list[dict[str, str]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [row for row in reader]


def _extract_rows_from_sheet(ws, sheet_name_requested: str | None) -> list[dict[str, str]]:
    """Read an openpyxl worksheet into a list of dicts keyed by header value.
    The unnamed column (empty header) after Notes On Job captures item descriptions
    from the shop's Customer Orders spreadsheet format.
    """
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h or "").strip() for h in header_row]

    # Track the index of the first unnamed column after col 5+ (item description column)
    item_desc_idx: int | None = None
    for i, h in enumerate(headers):
        if i >= 5 and h == "":
            item_desc_idx = i
            break

    rows = []
    for raw_row in ws.iter_rows(min_row=2, values_only=True):
        d: dict[str, str] = {}
        for i, v in enumerate(raw_row):
            if i >= len(headers):
                break
            key = headers[i] if headers[i] else f"__col{i}"
            d[key] = str(v or "").strip()
        # Expose the item description column with a stable key
        if item_desc_idx is not None:
            d["__item_desc"] = d.get(f"__col{item_desc_idx}", "")
        rows.append(d)
    return rows


@router.post("/import/sheets", response_model=list[str])
def list_import_sheets(
    file: UploadFile = File(...),
    _: AuthContext = Depends(get_auth_context),
):
    """Return the sheet names from an uploaded Excel file (CSV returns an empty list)."""
    content = file.file.read()
    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".xlsm", ".xls")):
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        return list(wb.sheetnames)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Could not read Excel file: {exc}")


@router.post("/import", response_model=CustomerOrderImportResult)
def import_customer_orders(
    file: UploadFile = File(...),
    dry_run: bool = Query(default=True),
    sheet_name: str | None = Query(default=None),
    session: Session = Depends(get_session),
    auth: AuthContext = Depends(get_auth_context),
    _: None = Depends(require_tech_or_above),
):
    content = file.file.read()

    rows: list[dict[str, str]] = []
    filename = (file.filename or "").lower()
    if filename.endswith((".xlsx", ".xlsm", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            rows = _extract_rows_from_sheet(ws, sheet_name)
        except Exception as exc:
            raise HTTPException(status_code=422, detail=f"Could not read Excel file: {exc}")
    else:
        try:
            rows = _read_csv_rows(content)
        except Exception as exc:
            raise HTTPException(status_code=422, detail=f"Could not parse CSV: {exc}")

    imported = 0
    skipped = 0
    skipped_reasons: dict[str, int] = {}

    def _skip(reason: str) -> None:
        nonlocal skipped
        skipped += 1
        skipped_reasons[reason] = skipped_reasons.get(reason, 0) + 1

    now = datetime.now(timezone.utc)
    customer_cache = _CustomerCache(session, auth.tenant_id) if not dry_run else None

    for row in rows:
        # ── Resolve title ─────────────────────────────────────────────────────
        # Priority: unnamed item-description column → Job column → title/item/description
        title = (
            row.get("__item_desc")
            or row.get("Job")
            or row.get("job")
            or row.get("title")
            or row.get("item")
            or row.get("description")
            or ""
        ).strip()
        if not title:
            _skip("missing item description")
            continue

        # ── Status ────────────────────────────────────────────────────────────
        gang = row.get("GA/NG") or row.get("ga/ng") or row.get("status") or ""
        ordered_date = row.get("Ordered date") or row.get("ordered_date") or row.get("ordered date") or ""
        status_raw = row.get("status", "")  # generic CSV fallback
        if gang or ordered_date:
            status, supplier_hint = _derive_status_and_supplier(gang, ordered_date)
        else:
            norm = _normalise(status_raw)
            status = norm if norm in _VALID_STATUSES else "to_order"
            supplier_hint = None

        # ── Priority ──────────────────────────────────────────────────────────
        priority_raw = _normalise(row.get("priority", "") or "normal")
        priority = priority_raw if priority_raw in _VALID_PRIORITIES else "normal"

        # ── Cost ──────────────────────────────────────────────────────────────
        cost_raw = (
            row.get("Quote") or row.get("quote")
            or row.get("cost") or row.get("estimated_cost") or row.get("price")
            or ""
        )
        estimated_cost_cents = _parse_cents(cost_raw)

        # ── Customer ──────────────────────────────────────────────────────────
        customer_name = (
            row.get("Customer Name") or row.get("customer_name")
            or row.get("customer") or ""
        ).strip()
        customer_phone = (
            row.get("Number") or row.get("number")
            or row.get("phone") or row.get("customer_phone") or ""
        ).strip()

        # ── Notes ─────────────────────────────────────────────────────────────
        notes = (
            row.get("Notes On Job") or row.get("notes_on_job")
            or row.get("notes") or row.get("note") or ""
        ).strip() or None

        # ── Supplier ──────────────────────────────────────────────────────────
        # Explicit supplier column takes precedence; fall back to hint from "Ordered date"
        supplier = (row.get("supplier") or row.get("Supplier") or "").strip() or supplier_hint or None

        customer_id: UUID | None = None
        if not dry_run and customer_cache is not None:
            customer_id = customer_cache.find_or_create(customer_name, customer_phone)
            order = CustomerOrder(
                tenant_id=auth.tenant_id,
                customer_id=customer_id,
                title=title,
                supplier=supplier,
                status=status,
                priority=priority,
                estimated_cost_cents=estimated_cost_cents,
                notes=notes,
                created_at=now,
                updated_at=now,
            )
            session.add(order)

        imported += 1

    if not dry_run:
        session.commit()

    return CustomerOrderImportResult(
        dry_run=dry_run,
        total_rows=len(rows),
        imported=imported,
        skipped=skipped,
        skipped_reasons=skipped_reasons,
    )
