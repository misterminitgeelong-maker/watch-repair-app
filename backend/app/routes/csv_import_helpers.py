"""Pure parsing / normalization helpers extracted from routes/csv_import.py.

These functions don't touch the database — they just turn uploaded CSV/XLSX
bytes into cleaned row dicts and infer statuses / numbers. They used to live
inline in csv_import.py (lines 58-520 of the 1200-line original). Keeping
them here makes csv_import.py focused on the three tenant-scoped pipelines
(watch / shoe / mobile) and makes the helpers unit-testable without having
to spin up the FastAPI app.
"""
from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import xlrd
from fastapi import HTTPException

def _parse_date(raw: str) -> date | None:
    if not raw or not raw.strip():
        return None
    raw = raw.strip()

    # Excel date serial (e.g. 44927 or 44927.0)
    try:
        n = float(raw)
        if 1000 < n < 100000 and n == int(n):
            from datetime import timedelta
            d = date(1899, 12, 30) + timedelta(days=int(n))
            return d
    except (ValueError, TypeError):
        pass

    for pattern, groups in [
        (r"^(\d{1,2})/(\d{1,2})/(\d{2})$", "dmy2"),
        (r"^(\d{1,2})/(\d{1,2})/(\d{4})$", "dmy4"),
        (r"^(\d{4})-(\d{1,2})-(\d{1,2})$", "ymd"),
        (r"^(\d{1,2})\.(\d{1,2})\.(\d{2,4})$", "dmy_dot"),
        (r"^(\d{1,2})/+(\d{1,2})/+(\d{2,4})$", "dmy_slash"),
    ]:
        m = re.match(pattern, raw)
        if not m:
            continue
        if groups == "ymd":
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        else:
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        if 1 <= mo <= 12 and 1 <= d <= 31:
            try:
                return date(y, mo, d)
            except ValueError:
                continue
    return None


def _parse_date_flexible(raw: str) -> date | None:
    """Like ``_parse_date`` but also tries the first whitespace-separated token (e.g. ``23.7.25 9.30am``)."""
    d = _parse_date(raw or "")
    if d:
        return d
    head = (raw or "").strip().split()[0] if (raw or "").strip() else ""
    return _parse_date(head) if head else None


def _normalize_phone(raw: str) -> str | None:
    if not raw or not raw.strip():
        return None
    digits = re.sub(r"\D", "", raw.strip())
    if len(digits) > 12:
        digits = digits[:10]
    if len(digits) == 9 and digits[0] in ("4", "3"):
        digits = "0" + digits
    if len(digits) < 8:
        return None
    return digits


# PostgreSQL INTEGER max; clamp to avoid overflow from corrupted Excel values
_MAX_CENTS = 2_147_483_647


def _dollars_to_cents(raw: str) -> int:
    if not raw or not raw.strip():
        return 0
    s = raw.strip().lower()
    if any(token in s for token in ["n/c", "nc", "quote"]):
        return 0

    # Handles values like "2x69.95" or "2 x $49.95 - 10%"
    m = re.search(r"(\d+)\s*x\s*\$?\s*(\d+(?:\.\d+)?)", s)
    if m:
        qty = int(m.group(1))
        unit = float(m.group(2))
        total = qty * unit
        discount = re.search(r"-\s*(\d+(?:\.\d+)?)\s*%", s)
        if discount:
            total *= 1 - (float(discount.group(1)) / 100)
        return min(int(round(total * 100)), _MAX_CENTS)

    numbers = [float(n) for n in re.findall(r"\d+(?:\.\d+)?", s)]
    if not numbers:
        return 0
    return min(int(round(numbers[0] * 100)), _MAX_CENTS)


def _clean_name(raw: str) -> str | None:
    if not raw or not raw.strip():
        return None
    name = re.sub(r"\s+", " ", raw.strip())
    lower_name = name.lower()

    # Common placeholders/comments that are not valid customer names.
    invalid_name_fragments = [
        "already collected",
        "has been collected",
        "picked up",
        "cancelled",
        "unknown",
    ]
    if any(fragment in lower_name for fragment in invalid_name_fragments):
        return None

    if re.match(r"^\d+$", name):
        return None
    if re.match(r"^\d{1,2}/\d{1,2}/\d{2,4}", name):
        return None
    if len(name) < 2:
        return None
    return name


def _get_first(row: dict[str, str], keys: list[str]) -> str:
    """Return the first non-empty value found in the provided column keys."""
    for key in keys:
        value = row.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _normalize_key(key: str) -> str:
    k = (key or "").strip().lower()
    k = re.sub(r"[^a-z0-9]+", "_", k)
    return k.strip("_")


def _normalize_row_keys(row: dict[str, str]) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for key, value in row.items():
        normalized[_normalize_key(key)] = (value or "").strip()

    # First column with no header (Excel: _col0→col0, CSV: "") often contains ticket/job numbers.
    for blank_key in ("", "_col0", "col0"):
        val = (row.get(blank_key) or "").strip()
        if val and "original_job_id" not in normalized:
            try:
                if re.match(r"^[\d.]+$", val):
                    normalized["original_job_id"] = val
                    break
            except Exception:
                pass
    return normalized


def _to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _load_csv_rows(raw_bytes: bytes) -> list[dict[str, str]]:
    text = ""
    for encoding in ("utf-8-sig", "utf-16", "latin-1"):
        try:
            text = raw_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue

    if not text:
        raise HTTPException(status_code=400, detail="Unable to decode CSV file. Please export as UTF-8 CSV.")

    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    except csv.Error:
        reader = csv.DictReader(io.StringIO(text))

    rows = list(reader)
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV header row could not be read. Ensure the first row contains column names.")
    return rows


def _score_repair_header_row(header_row: tuple) -> int:
    """Prefer worksheets that look like a repair/job log (flexible column names)."""
    keys = {_normalize_key(_to_text(v) or f"_col{i}") for i, v in enumerate(header_row)}
    score = 0
    if {"name", "customer", "client_name", "customer_name"} & keys:
        score += 2
    if {"number", "phone", "phone_number", "mobile"} & keys:
        score += 2
    if {"brand", "brand_case_numbers", "make_model", "model"} & keys:
        score += 1
    if {"date", "date_in", "date_time", "datetime", "created_at", "date_recieved", "date_received"} & keys:
        score += 1
    if "quote" in keys or "quote_price" in keys:
        score += 1
    if {"repair_notes", "notes", "notes_on_job", "description", "job_description", "other_notes"} & keys:
        score += 1
    if {"status", "ga_ng_collected", "job_status"} & keys:
        score += 1
    if {"address", "job_address"} & keys:
        score += 1
    if {"docket_number", "ticket", "ticket_number", "col0", "_col0"} & keys or any(
        k in keys for k in keys if k.startswith("col") and k[3:].isdigit()
    ):
        score += 1
    return score


def _pick_xlsx_sheet(workbook, sheet_name: str | None):
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            names = ", ".join(workbook.sheetnames[:12])
            more = f" (+{len(workbook.sheetnames) - 12} more)" if len(workbook.sheetnames) > 12 else ""
            raise HTTPException(
                status_code=400,
                detail=f'Worksheet "{sheet_name}" not found. Available: {names}{more}',
            )
        return workbook[sheet_name]
    best_ws = None
    best_score = -1
    for ws in workbook.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        sc = _score_repair_header_row(rows[0])
        if sc > best_score:
            best_score = sc
            best_ws = ws
    if best_ws is not None and best_score >= 3:
        return best_ws
    return workbook.active


def _load_xlsx_rows(raw_bytes: bytes, sheet_name: str | None = None) -> tuple[list[dict[str, str]], str]:
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
        sheet = _pick_xlsx_sheet(workbook, sheet_name)
        if sheet is None:
            raise HTTPException(status_code=400, detail="Excel file has no worksheets.")
        all_rows = list(sheet.iter_rows(values_only=True))
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to read uploaded Excel file")
        raise HTTPException(
            status_code=400,
            detail="Could not read the Excel file. Try saving as CSV (UTF-8) or a fresh .xlsx.",
        ) from exc

    if not all_rows:
        return [], sheet.title

    header_row = all_rows[0]
    header = [_to_text(v) or f"_col{i}" for i, v in enumerate(header_row)]
    # Ensure unique header keys (Excel often has blank columns)
    seen: set[str] = set()
    for i, h in enumerate(header):
        if h in seen:
            header[i] = f"{h}_{i}"
        seen.add(header[i])

    result: list[dict[str, str]] = []
    for row in all_rows[1:]:
        if not any(_to_text(c) for c in row):
            continue
        row_len = len(row)
        hdr_len = len(header)
        cells = [_to_text(row[idx]) if idx < row_len else "" for idx in range(hdr_len)]
        result.append({header[i]: cells[i] for i in range(hdr_len)})
    return result, sheet.title


def _load_xls_rows(raw_bytes: bytes) -> list[dict[str, str]]:
    workbook = xlrd.open_workbook(file_contents=raw_bytes)
    sheet = workbook.sheet_by_index(0)
    if sheet.nrows == 0:
        return []

    header = [_to_text(sheet.cell_value(0, col)) for col in range(sheet.ncols)]
    output: list[dict[str, str]] = []
    for row_idx in range(1, sheet.nrows):
        row_values = [_to_text(sheet.cell_value(row_idx, col)) for col in range(sheet.ncols)]
        if not any(row_values):
            continue
        output.append({header[idx]: value for idx, value in enumerate(row_values)})
    return output


def _load_rows(
    filename: str, raw_bytes: bytes, sheet_name: str | None = None
) -> tuple[list[dict[str, str]], str, str | None]:
    ext = Path(filename).suffix.lower()
    if ext == ".csv":
        return _load_csv_rows(raw_bytes), "csv", None
    if ext in {".xlsx", ".xlsm"}:
        rows, title = _load_xlsx_rows(raw_bytes, sheet_name)
        return rows, "xlsx", title
    if ext == ".xls":
        return _load_xls_rows(raw_bytes), "xls", None

    # Be tolerant of renamed/missing extensions by sniffing common formats.
    if raw_bytes.startswith(b"PK"):
        rows, title = _load_xlsx_rows(raw_bytes, sheet_name)
        return rows, "xlsx", title
    if raw_bytes.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"):
        return _load_xls_rows(raw_bytes), "xls", None

    try:
        return _load_csv_rows(raw_bytes), "csv", None
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail="Only .csv, .xlsx, .xls, and .xlsm files are accepted. If this is CSV, re-save as UTF-8 CSV.",
        ) from exc


_STATUS_MAP = {
    "collected": "collected",
    "done": "completed",
    "ga": "completed",  # Good as New
    "ng": "no_go",     # No Good
    "in_repair": "working_on",
    "in repair": "working_on",
    "working on": "working_on",
    "ready": "awaiting_collection",
    "awaiting collection": "awaiting_collection",
    "intake": "awaiting_go_ahead",
    "awaiting go ahead": "awaiting_go_ahead",
    "go ahead": "go_ahead",
    "no go": "no_go",
    "completed": "completed",
    "awaiting parts": "awaiting_parts",
    "parts to be ordered": "parts_to_order",
    "sent to labanda": "sent_to_labanda",
    "service": "service",
    "cancelled": "no_go",
    "approved": "go_ahead",
}

# Status cells like "Sent Back Completed" → slug sent_back_completed
_SLUG_STATUS_ALIASES: dict[str, str] = {
    "sent_back_completed": "completed",
    "ready_for_collection": "awaiting_collection",
}


def _status_slug(status_raw: str) -> str:
    s = (status_raw or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return re.sub(r"_+", "_", s).strip("_")


def _allocate_import_job_number(usage: dict[str, int], ticket_stem: str | None, job_seq: int) -> str:
    """Assign IMP-{ticket} or IMP-{seq}; duplicate tickets become IMP-{ticket}-2, -3, …"""
    if ticket_stem:
        base = f"IMP-{ticket_stem}"
    else:
        base = f"IMP-{job_seq:05d}"
    n = usage.get(base, 0)
    usage[base] = n + 1
    if n == 0:
        return base
    return f"{base}-{n + 1}"


def _reserve_import_job_number(preferred: str, used_job_numbers: set[str]) -> str:
    """
    Pick a job_number not in ``used_job_numbers``, starting from preferred (in-file slot).
    Updates ``used_job_numbers`` with the chosen value (in-memory; avoids per-row DB round-trips).
    """
    m = re.match(r"^(IMP-\d+)(?:-(\d+))?$", preferred)
    if m:
        root = m.group(1)
        k = int(m.group(2)) if m.group(2) else 1
    else:
        root = preferred
        k = 1
    while True:
        candidate = root if k == 1 else f"{root}-{k}"
        if candidate not in used_job_numbers:
            used_job_numbers.add(candidate)
            return candidate
        k += 1


def _infer_job_status(status_raw: str, notes_raw: str) -> str:
    status = (status_raw or "").strip().lower()
    notes = (notes_raw or "").strip().lower()
    slug = _status_slug(status_raw)

    if slug in _SLUG_STATUS_ALIASES:
        return _SLUG_STATUS_ALIASES[slug]
    if slug in _STATUS_MAP:
        return _STATUS_MAP[slug]
    if status in _STATUS_MAP:
        return _STATUS_MAP[status]

    if status == "delivered":
        if any(token in notes for token in ["collected", "picked up", "has been collected", "watch collect"]):
            return "collected"
        if any(token in notes for token in ["awaiting collection", "ready for collection", "ready in drawer", "messaged for collection"]):
            return "awaiting_collection"
        return "completed"

    return "awaiting_go_ahead"


def _infer_auto_key_status(status_raw: str, notes_raw: str) -> str:
    """Map spreadsheet status to AutoKeyJob.status (JobStatus literals)."""
    slug = _status_slug(status_raw)
    s = (status_raw or "").strip().lower()

    mobile_slug_map: dict[str, str] = {
        "booked": "booked",
        "confirmed": "booked",
        "scheduled": "booked",
        "pending_booking": "pending_booking",
        "awaiting_booking": "pending_booking",
        "en_route": "en_route",
        "on_the_way": "en_route",
        "driving": "en_route",
        "on_site": "on_site",
        "arrived": "on_site",
        "at_site": "on_site",
        "awaiting_customer_details": "awaiting_customer_details",
    }
    if slug in mobile_slug_map:
        return mobile_slug_map[slug]
    if s in mobile_slug_map:
        return mobile_slug_map[s]

    base = _infer_job_status(status_raw, notes_raw)
    if base == "awaiting_go_ahead" and not (status_raw or "").strip():
        return "awaiting_quote"
    if base in {
        "awaiting_go_ahead",
        "go_ahead",
        "working_on",
        "awaiting_parts",
        "parts_to_order",
        "sent_to_labanda",
        "quoted_by_labanda",
        "service",
        "completed",
        "awaiting_collection",
        "collected",
        "no_go",
        "en_route",
        "on_site",
        "pending_booking",
        "booked",
        "awaiting_customer_details",
    }:
        return base
    return "awaiting_quote"
