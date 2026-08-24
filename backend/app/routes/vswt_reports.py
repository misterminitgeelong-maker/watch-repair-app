"""VSWT Regional Intelligence.

Ingests Mister Minit HQ's weekly "VSWT-WSS" regional Excel export (one row per shop, covering
every shop in the VSWT region) into a shared table, then serves each logged-in shop its own
rank against the rest of the region — and against a "peer group" of comparable franchise stores —
across every KPI HQ tracks.

Not tenant-scoped like most Mainspring tables: this is shared regional reference data. A shop's
own row is found via `Tenant.shop_number` (already used elsewhere for Minit shop identity), so
there is no picker — whatever shop is logged in sees its own numbers automatically. Endpoints
return `{"available": False, ...}` rather than 404 when the logged-in tenant has no shop_number,
or when that shop_number isn't present in the region's data yet — the frontend uses this to decide
whether to show the section at all, the same way the Reports page already hides the shoe-repair
section when there's nothing to show.

See app/vswt_kpis.py for the column layout / KPI list this parser and these rankings are built
from — keep field names in sync with that module rather than re-deriving them here.
"""
from __future__ import annotations

import io
import statistics
from datetime import date, datetime, timezone
from typing import Any, Literal, Optional

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.concurrency import run_in_threadpool
from fastapi.responses import Response
from sqlmodel import Session, delete as sa_delete, select

from ..database import get_session
from ..dependencies import AuthContext, get_auth_context, require_manager_or_above
from ..models import Tenant, VswtWeeklyShopMetric
from ..pdf_vswt_report import build_weekly_report_pdf
from ..vswt_kpis import (
    CATEGORY_SALES_KEYS,
    COLUMN_MAP,
    KPI_DEFS,
    KPI_GROUPS,
    KpiDef,
    PEER_COMP_STATUS,
    PEER_FORMAT,
    clean_cell,
)
from sqlmodel import SQLModel

router = APIRouter(prefix="/v1/reports/vswt", tags=["vswt"])

_DATA_START_ROW = 6  # 1-based Excel row; shop rows start here
_WEEK_NUMBER_ROW = 3  # 1-based Excel row: "Week Number:" label in col A, value in col C
_WEEK_NUMBER_COL = 3  # column C


# ── Parsing (Summary sheet -> list[dict]) ───────────────────────────────────────────────────

def _parse_vswt_workbook(raw_bytes: bytes, filename: str) -> dict[str, Any]:
    # read_only=True makes openpyxl stream the sheet instead of building its full in-memory
    # object model (styles, formatting, merged cells, etc.) — we only ever read cell values via
    # iter_rows(), so this is a large speed/memory win on real HQ export files, which carry much
    # more formatting than the column data we actually use. Without it, a handful of files
    # uploaded together was slow enough (CPU-bound, single-worker deployment) to trip the
    # platform's health check and get the request killed mid-response.
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=400, detail=f"Could not read '{filename}' as an Excel file: {exc}"
        ) from exc

    try:
        sheet = None
        for name in wb.sheetnames:
            if name.strip().lower() == "summary":
                sheet = wb[name]
                break
        if sheet is None:
            sheet = wb.active
        grid = list(sheet.iter_rows(values_only=True))
    finally:
        wb.close()

    internal_week_raw = None
    if len(grid) >= _WEEK_NUMBER_ROW:
        row = grid[_WEEK_NUMBER_ROW - 1]
        if row and len(row) >= _WEEK_NUMBER_COL:
            internal_week_raw = clean_cell(row[_WEEK_NUMBER_COL - 1])
    internal_week: Optional[int] = None
    if isinstance(internal_week_raw, (int, float)):
        internal_week = int(internal_week_raw)

    rows: list[dict[str, Any]] = []
    for r in grid[_DATA_START_ROW - 1:]:
        if not r or len(r) < 2:
            continue
        shop_number = clean_cell(r[1])  # column B
        if shop_number in (None, ""):
            continue
        row: dict[str, Any] = {}
        for field, col in COLUMN_MAP:
            value = clean_cell(r[col - 1]) if len(r) >= col else None
            row[field] = value
        sn = row.get("shop_number")
        if isinstance(sn, float) and sn.is_integer():
            row["shop_number"] = str(int(sn))
        elif sn is not None:
            row["shop_number"] = str(sn).strip()
        rows.append(row)

    return {
        "filename": filename,
        "internal_week": internal_week,
        "rows": rows,
        "shop_count": len(rows),
    }


# ── Ranking helpers (ported from the reference app's rankOf/average/peerRows) ──────────────

def _rank_of(week_rows: list[VswtWeeklyShopMetric], key: str, value: Optional[float]) -> Optional[int]:
    if value is None:
        return None
    higher = sum(1 for r in week_rows if getattr(r, key) is not None and getattr(r, key) > value)
    return higher + 1


def _average(values: list[Optional[float]]) -> Optional[float]:
    vals = [v for v in values if v is not None]
    if not vals:
        return None
    return sum(vals) / len(vals)


def _peer_rows(week_rows: list[VswtWeeklyShopMetric]) -> list[VswtWeeklyShopMetric]:
    return [r for r in week_rows if r.store_format == PEER_FORMAT and r.comp_status == PEER_COMP_STATUS]


def _find_shop(week_rows: list[VswtWeeklyShopMetric], shop_number: str) -> Optional[VswtWeeklyShopMetric]:
    return next((r for r in week_rows if r.shop_number == shop_number), None)


def _ranks_for_week(week_rows: list[VswtWeeklyShopMetric], key: str) -> dict[str, int]:
    """Rank of every shop with a non-null `key` value for one week, all at once — O(n log n)
    instead of calling `_rank_of` once per shop (which is itself O(n), so O(n^2) over a whole
    week). Same "1 + count of strictly-greater values" tie semantics as `_rank_of`; used where we
    need every shop's rank for a week, not just one shop's (e.g. consistency leaderboards)."""
    present = sorted(
        ((r.shop_number, getattr(r, key)) for r in week_rows if getattr(r, key) is not None),
        key=lambda pair: pair[1],
        reverse=True,
    )
    ranks: dict[str, int] = {}
    for i, (shop_number, value) in enumerate(present):
        prev_shop, prev_value = present[i - 1] if i > 0 else (None, None)
        ranks[shop_number] = ranks[prev_shop] if value == prev_value else i + 1
    return ranks


def _rank_in_averages(averages: dict[str, float], shop_number: str) -> Optional[int]:
    """Same ranking rule as `_rank_of`, but over a plain {shop_number: average_value} dict rather
    than ORM rows — for ranking a shop's Month/Year average against the rest of the region's."""
    value = averages.get(shop_number)
    if value is None:
        return None
    higher = sum(1 for v in averages.values() if v > value)
    return higher + 1


def _all_weeks(session: Session) -> list[int]:
    return list(
        session.exec(
            select(VswtWeeklyShopMetric.week_seq).distinct().order_by(VswtWeeklyShopMetric.week_seq)
        ).all()
    )


def _week_rows(session: Session, week: int) -> list[VswtWeeklyShopMetric]:
    return list(
        session.exec(select(VswtWeeklyShopMetric).where(VswtWeeklyShopMetric.week_seq == week)).all()
    )


def _shop_number_for(auth: AuthContext, session: Session) -> Optional[str]:
    tenant = session.get(Tenant, auth.tenant_id)
    if not tenant or not tenant.shop_number:
        return None
    return tenant.shop_number


# ── Upload / commit ──────────────────────────────────────────────────────────────────────

@router.post("/upload")
async def upload_vswt_files(
    files: list[UploadFile] = File(...),
    auth: AuthContext = Depends(require_manager_or_above),
    session: Session = Depends(get_session),
):
    if not files:
        raise HTTPException(status_code=400, detail="No files provided.")

    parsed: list[dict[str, Any]] = []
    failed: list[str] = []
    for f in files:
        if not f.filename:
            continue
        raw = await f.read()
        try:
            result = await run_in_threadpool(_parse_vswt_workbook, raw, f.filename)
        except HTTPException:
            failed.append(f.filename)
            continue
        if not result["rows"]:
            failed.append(f.filename)
            continue
        parsed.append(result)

    if not parsed:
        raise HTTPException(
            status_code=400,
            detail=(
                "Couldn't read any shop rows from the selected file(s) — make sure they're "
                "VSWT-WSS files with a Summary tab."
            ),
        )

    existing_weeks = set(_all_weeks(session))
    used = set(existing_weeks)
    # Assign week numbers: prefer each file's own detected week if free, else next free number.
    parsed.sort(key=lambda p: (p["internal_week"] is None, p["internal_week"]))
    next_free = (max(used) + 1) if used else 1
    batch = []
    for p in parsed:
        week_number = p["internal_week"]
        if week_number is None or week_number in used:
            while next_free in used:
                next_free += 1
            week_number = next_free
        used.add(week_number)
        next_free = max(next_free, week_number + 1)
        batch.append(
            {
                "filename": p["filename"],
                "internal_week": p["internal_week"],
                "week_number": week_number,
                "shop_count": p["shop_count"],
                "overwrite": week_number in existing_weeks,
                "rows": p["rows"],
            }
        )

    # Auto-assignment always steers clear of collisions (see loop above), so a per-file
    # "overwrite" flag is only ever true here in the degenerate case where two source files
    # detected the exact same week and neither could be bumped — the live case that actually
    # matters is the user *editing* the week number in the confirm step back onto an existing
    # week, which the frontend detects itself by checking the edited number against this list.
    return {"failed_files": failed, "batch": batch, "existing_weeks": sorted(existing_weeks)}


class VswtCommitFile(SQLModel):
    filename: str
    week_number: int
    rows: list[dict[str, Any]]


class VswtCommitRequest(SQLModel):
    batch: list[VswtCommitFile]


@router.post("/commit")
def commit_vswt_batch(
    payload: VswtCommitRequest,
    auth: AuthContext = Depends(require_manager_or_above),
    session: Session = Depends(get_session),
):
    if not payload.batch:
        raise HTTPException(status_code=400, detail="Nothing to commit.")
    week_numbers = [item.week_number for item in payload.batch]
    if len(set(week_numbers)) != len(week_numbers):
        raise HTTPException(status_code=400, detail="Duplicate week numbers in this batch.")

    metric_fields = [name for name, _ in COLUMN_MAP]
    now = datetime.now(timezone.utc)
    saved = []
    for item in payload.batch:
        session.exec(
            sa_delete(VswtWeeklyShopMetric).where(VswtWeeklyShopMetric.week_seq == item.week_number)
        )
        count = 0
        for row in item.rows:
            if not row.get("shop_number"):
                continue
            fields = {name: row.get(name) for name in metric_fields}
            session.add(
                VswtWeeklyShopMetric(
                    week_seq=item.week_number,
                    source_filename=item.filename,
                    uploaded_by_tenant_id=auth.tenant_id,
                    uploaded_by_user_id=auth.user_id,
                    uploaded_at=now,
                    **fields,
                )
            )
            count += 1
        saved.append({"week_number": item.week_number, "shop_count": count})
    session.commit()
    return {"saved": saved}


@router.get("/weeks")
def get_vswt_weeks(
    auth: AuthContext = Depends(require_manager_or_above),
    session: Session = Depends(get_session),
):
    weeks = _all_weeks(session)
    out = []
    for w in weeks:
        rows = _week_rows(session, w)
        source_filenames = sorted({r.source_filename for r in rows if r.source_filename})
        uploaded_at = max((r.uploaded_at for r in rows), default=None)
        out.append(
            {
                "week": w,
                "shop_count": len(rows),
                "source_filenames": source_filenames,
                "uploaded_at": uploaded_at,
            }
        )
    return {"weeks": out}


@router.delete("/weeks/{week_seq}")
def delete_vswt_week(
    week_seq: int,
    auth: AuthContext = Depends(require_manager_or_above),
    session: Session = Depends(get_session),
):
    session.exec(sa_delete(VswtWeeklyShopMetric).where(VswtWeeklyShopMetric.week_seq == week_seq))
    session.commit()
    return {"deleted_week": week_seq}


# ── Read endpoints ───────────────────────────────────────────────────────────────────────

@router.get("/summary")
def get_vswt_summary(
    auth: AuthContext = Depends(get_auth_context),
    session: Session = Depends(get_session),
):
    shop_number = _shop_number_for(auth, session)
    if shop_number is None:
        return {"available": False, "reason": "no_shop_number"}

    weeks = _all_weeks(session)
    if not weeks:
        return {"available": False, "reason": "no_data"}

    latest = weeks[-1]
    prev = weeks[-2] if len(weeks) > 1 else None
    latest_rows = _week_rows(session, latest)
    my_row = _find_shop(latest_rows, shop_number)
    if my_row is None:
        return {"available": False, "reason": "shop_not_found", "latest_week": latest}

    prev_row = _find_shop(_week_rows(session, prev), shop_number) if prev is not None else None
    peers = _peer_rows(latest_rows)
    area_rows = [r for r in latest_rows if my_row.area_name and r.area_name == my_row.area_name]

    return {
        "available": True,
        "shop_number": shop_number,
        "shop_name": my_row.shop_name,
        "area_name": my_row.area_name,
        "latest_week": latest,
        "weeks_tracked": len(weeks),
        "region_size": len(latest_rows),
        "peer_size": len(peers),
        "area_size": len(area_rows),
        "sales": {
            "value": my_row.sales_ty,
            "prev_value": prev_row.sales_ty if prev_row else None,
            "region_rank": _rank_of(latest_rows, "sales_ty", my_row.sales_ty),
            "peer_rank": _rank_of(peers, "sales_ty", my_row.sales_ty),
            "area_rank": _rank_of(area_rows, "sales_ty", my_row.sales_ty) if area_rows else None,
        },
        "customers": {
            "value": my_row.customer_ty,
            "prev_value": prev_row.customer_ty if prev_row else None,
            "region_rank": _rank_of(latest_rows, "customer_ty", my_row.customer_ty),
        },
        "jobs": {
            "value": my_row.jobs_ty,
            "prev_value": prev_row.jobs_ty if prev_row else None,
            "region_rank": _rank_of(latest_rows, "jobs_ty", my_row.jobs_ty),
        },
    }


@router.get("/scorecard")
def get_vswt_scorecard(
    shop_number: Optional[str] = Query(
        None, description="View another Minit shop's scorecard instead of your own (you must be a Minit shop yourself)."
    ),
    auth: AuthContext = Depends(get_auth_context),
    session: Session = Depends(get_session),
):
    my_shop_number = _shop_number_for(auth, session)
    if my_shop_number is None:
        return {"available": False, "reason": "no_shop_number"}
    target_shop_number = shop_number or my_shop_number
    weeks = _all_weeks(session)
    if not weeks:
        return {"available": False, "reason": "no_data"}

    matrix = []
    found_any = False
    target_name = None
    target_area = None
    for w in weeks:
        week_rows = _week_rows(session, w)
        target_row = _find_shop(week_rows, target_shop_number)
        if target_row is not None:
            found_any = True
            target_name = target_row.shop_name
            target_area = target_row.area_name
        cells = {}
        for kpi in KPI_DEFS:
            value = getattr(target_row, kpi.key) if target_row else None
            cells[kpi.key] = {"value": value, "rank": _rank_of(week_rows, kpi.key, value)}
        matrix.append({"week": w, "region_size": len(week_rows), "cells": cells})

    if not found_any:
        return {"available": False, "reason": "shop_not_found"}

    return {
        "available": True,
        "shop_number": target_shop_number,
        "shop_name": target_name,
        "area_name": target_area,
        "viewing_own_shop": target_shop_number == my_shop_number,
        "weeks": weeks,
        "groups": KPI_GROUPS,
        "kpis": [{"key": k.key, "label": k.label, "group": k.group, "type": k.type} for k in KPI_DEFS],
        "matrix": matrix,
    }


@router.get("/rankings")
def get_vswt_rankings(
    week: Optional[int] = Query(None),
    shop_number: Optional[str] = Query(
        None, description="View another Minit shop's rankings instead of your own (you must be a Minit shop yourself)."
    ),
    auth: AuthContext = Depends(get_auth_context),
    session: Session = Depends(get_session),
):
    my_shop_number = _shop_number_for(auth, session)
    if my_shop_number is None:
        return {"available": False, "reason": "no_shop_number"}
    target_shop_number = shop_number or my_shop_number
    weeks = _all_weeks(session)
    if not weeks:
        return {"available": False, "reason": "no_data"}
    target_week = week if week in weeks else weeks[-1]

    week_rows = _week_rows(session, target_week)
    target_row = _find_shop(week_rows, target_shop_number)
    if target_row is None:
        return {"available": False, "reason": "shop_not_found", "week": target_week}

    peers = _peer_rows(week_rows)
    rows = []
    for kpi in KPI_DEFS:
        value = getattr(target_row, kpi.key)
        region_rank = _rank_of(week_rows, kpi.key, value)
        percentile = (
            (len(week_rows) - region_rank) / (len(week_rows) - 1)
            if region_rank is not None and len(week_rows) > 1
            else None
        )
        rows.append(
            {
                "key": kpi.key,
                "label": kpi.label,
                "group": kpi.group,
                "type": kpi.type,
                "value": value,
                "region_avg": _average([getattr(r, kpi.key) for r in week_rows]),
                "region_rank": region_rank,
                "percentile": percentile,
                "peer_avg": _average([getattr(r, kpi.key) for r in peers]),
                "peer_rank": _rank_of(peers, kpi.key, value),
            }
        )

    return {
        "available": True,
        "shop_number": target_shop_number,
        "shop_name": target_row.shop_name,
        "area_name": target_row.area_name,
        "viewing_own_shop": target_shop_number == my_shop_number,
        "week": target_week,
        "weeks": weeks,
        "region_size": len(week_rows),
        "peer_size": len(peers),
        "rows": rows,
    }


@router.get("/directory")
def get_vswt_directory(
    week: Optional[int] = Query(None),
    search: Optional[str] = Query(None, description="Filter by shop name, shop number, or area (case-insensitive)."),
    group: Optional[str] = Query(None, description="KPI group to include as columns; defaults to Headline."),
    peer_only: bool = Query(False, description="Only franchise + comparable stores."),
    auth: AuthContext = Depends(get_auth_context),
    session: Session = Depends(get_session),
):
    """Every shop in the region for one week, searchable — the entry point for browsing/looking
    up any other Minit shop's numbers, not just your own. Gated the same as the rest of this
    feature: you must be a Minit shop yourself to browse the network."""
    my_shop_number = _shop_number_for(auth, session)
    if my_shop_number is None:
        return {"available": False, "reason": "no_shop_number"}
    weeks = _all_weeks(session)
    if not weeks:
        return {"available": False, "reason": "no_data"}
    target_week = week if week in weeks else weeks[-1]

    week_rows = _week_rows(session, target_week)
    peer_numbers = {r.shop_number for r in _peer_rows(week_rows)}

    kpi_group = group if group in KPI_GROUPS else "Headline"
    kpis = [k for k in KPI_DEFS if k.group == kpi_group]

    rows = week_rows
    if peer_only:
        rows = [r for r in rows if r.shop_number in peer_numbers]
    if search and search.strip():
        q = search.strip().lower()
        rows = [
            r for r in rows
            if q in (r.shop_name or "").lower()
            or q in (r.shop_number or "").lower()
            or q in (r.area_name or "").lower()
        ]

    out_rows = [
        {
            "shop_number": r.shop_number,
            "shop_name": r.shop_name,
            "area_name": r.area_name,
            "store_format": r.store_format,
            "comp_status": r.comp_status,
            "is_peer": r.shop_number in peer_numbers,
            "is_me": r.shop_number == my_shop_number,
            "values": {k.key: getattr(r, k.key) for k in kpis},
        }
        for r in rows
    ]

    return {
        "available": True,
        "week": target_week,
        "weeks": weeks,
        "region_size": len(week_rows),
        "peer_size": len(peer_numbers),
        "result_size": len(out_rows),
        "group": kpi_group,
        "groups": KPI_GROUPS,
        "kpis": [{"key": k.key, "label": k.label, "group": k.group, "type": k.type} for k in kpis],
        "rows": out_rows,
    }


@router.get("/shop-report")
def get_vswt_shop_report(
    shop_number: Optional[str] = Query(
        None, description="View another Minit shop's report instead of your own (you must be a Minit shop yourself)."
    ),
    group: Optional[str] = Query(None, description="KPI group to include as rows; defaults to Headline."),
    auth: AuthContext = Depends(get_auth_context),
    session: Session = Depends(get_session),
):
    """One shop's Week / Month / Year numbers, each ranked against the rest of the region — the
    Directory's per-shop drill-down. "Month" is a rolling trailing-4-week average; "Year" averages
    every week on file (there's no calendar date on this data, only a sequence of weekly uploads,
    so "Year" grows into a real year as more weeks accumulate). Averaging over many weeks — rather
    than only ever showing the latest week — is what surfaces a shop that's consistently strong or
    weak, instead of one that just had a single standout or disastrous week.
    """
    my_shop_number = _shop_number_for(auth, session)
    if my_shop_number is None:
        return {"available": False, "reason": "no_shop_number"}
    target_shop_number = shop_number or my_shop_number
    weeks = _all_weeks(session)
    if not weeks:
        return {"available": False, "reason": "no_data"}

    rows_by_week = {w: _week_rows(session, w) for w in weeks}
    latest_week = weeks[-1]
    latest_rows = rows_by_week[latest_week]
    target_latest = _find_shop(latest_rows, target_shop_number)
    if target_latest is None:
        return {"available": False, "reason": "shop_not_found"}

    month_weeks = set(weeks[-4:])
    kpi_group = group if group in KPI_GROUPS else "Headline"
    kpis = [k for k in KPI_DEFS if k.group == kpi_group]

    # Single pass over every week: accumulate each shop's raw values for the Month window and for
    # all-time ("Year"), and record the target shop's own rank each week it appears (used below
    # for the Year consistency stats — best/worst rank and how much it's varied).
    month_raw: dict[str, dict[str, list[float]]] = {}  # shop_number -> kpi.key -> values
    year_raw: dict[str, dict[str, list[float]]] = {}
    target_week_ranks: dict[str, list[int]] = {k.key: [] for k in kpis}
    for w in weeks:
        week_rows = rows_by_week[w]
        target_row = _find_shop(week_rows, target_shop_number)
        if target_row is not None:
            for kpi in kpis:
                rank = _rank_of(week_rows, kpi.key, getattr(target_row, kpi.key))
                if rank is not None:
                    target_week_ranks[kpi.key].append(rank)
        in_month = w in month_weeks
        for r in week_rows:
            for kpi in kpis:
                v = getattr(r, kpi.key)
                if v is None:
                    continue
                year_raw.setdefault(r.shop_number, {}).setdefault(kpi.key, []).append(v)
                if in_month:
                    month_raw.setdefault(r.shop_number, {}).setdefault(kpi.key, []).append(v)

    # Region-wide average-per-shop, per KPI — used to rank the target's Month/Year average against
    # everyone else's, the same way `_rank_of` ranks a single week's value.
    month_avgs = {
        kpi.key: {sn: _average(vals[kpi.key]) for sn, vals in month_raw.items() if vals.get(kpi.key)}
        for kpi in kpis
    }
    year_avgs = {
        kpi.key: {sn: _average(vals[kpi.key]) for sn, vals in year_raw.items() if vals.get(kpi.key)}
        for kpi in kpis
    }

    rows = []
    for kpi in kpis:
        week_value = getattr(target_latest, kpi.key)
        month_weeks_counted = len(month_raw.get(target_shop_number, {}).get(kpi.key, []))
        year_weeks_counted = len(year_raw.get(target_shop_number, {}).get(kpi.key, []))
        week_ranks = target_week_ranks[kpi.key]
        rows.append(
            {
                "key": kpi.key, "label": kpi.label, "group": kpi.group, "type": kpi.type,
                "week": {"value": week_value, "rank": _rank_of(latest_rows, kpi.key, week_value)},
                "month": {
                    "value": month_avgs[kpi.key].get(target_shop_number),
                    "rank": _rank_in_averages(month_avgs[kpi.key], target_shop_number),
                    "weeks_counted": month_weeks_counted,
                },
                "year": {
                    "value": year_avgs[kpi.key].get(target_shop_number),
                    "rank": _rank_in_averages(year_avgs[kpi.key], target_shop_number),
                    "weeks_counted": year_weeks_counted,
                    "best_rank": min(week_ranks) if week_ranks else None,
                    "worst_rank": max(week_ranks) if week_ranks else None,
                    "rank_stdev": statistics.pstdev(week_ranks) if len(week_ranks) > 1 else None,
                },
            }
        )

    return {
        "available": True,
        "shop_number": target_shop_number,
        "shop_name": target_latest.shop_name,
        "area_name": target_latest.area_name,
        "viewing_own_shop": target_shop_number == my_shop_number,
        "weeks_tracked": len(weeks),
        "region_size": len(latest_rows),
        "group": kpi_group,
        "groups": KPI_GROUPS,
        "kpis": [{"key": k.key, "label": k.label, "group": k.group, "type": k.type} for k in kpis],
        "rows": rows,
    }


def _consistency_boards(
    session: Session, weeks: list[int], kpis: list[KpiDef], shop_number: Optional[str]
) -> list[dict[str, Any]]:
    """Top/bottom leaderboards ranked by average rank across every week on file, instead of one
    week's value — rewards a shop that's reliably strong over a shop that just had one great week.
    Mirrors the shape of the "latest" leaderboards below exactly, so the frontend can render both
    with the same component."""
    rank_lists: dict[str, dict[str, list[int]]] = {}
    shop_meta: dict[str, tuple[Optional[str], Optional[str]]] = {}
    for w in weeks:
        week_rows = _week_rows(session, w)
        for r in week_rows:
            shop_meta[r.shop_number] = (r.shop_name, r.area_name)
        for kpi in kpis:
            for sn, rank in _ranks_for_week(week_rows, kpi.key).items():
                rank_lists.setdefault(sn, {}).setdefault(kpi.key, []).append(rank)

    # A shop needs at least a handful of weeks on file before its average rank means anything —
    # otherwise one lucky week could pass for "consistency". Caps at 3 so this doesn't lock the
    # feature out entirely while only a few weeks have been uploaded so far.
    min_weeks = min(3, len(weeks))

    boards = []
    for kpi in kpis:
        entries = [
            (sn, sum(ranks) / len(ranks), len(ranks))
            for sn, by_kpi in rank_lists.items()
            for ranks in [by_kpi.get(kpi.key, [])]
            if len(ranks) >= min_weeks
        ]
        entries.sort(key=lambda e: e[1])  # ascending average rank = most consistently good first
        total = len(entries)
        my_index = next((i for i, e in enumerate(entries) if e[0] == shop_number), None)

        def _entry(rank_pos: int, entry: tuple[str, float, int]) -> dict[str, Any]:
            sn, avg_rank, weeks_counted = entry
            is_me = sn == shop_number
            name, _area = shop_meta.get(sn, (None, None))
            return {
                "rank": rank_pos,
                "shop_number": sn,
                "shop_name": name,
                "value": round(avg_rank, 2),
                "weeks_counted": weeks_counted,
                "is_me": is_me,
            }

        top = [_entry(i + 1, e) for i, e in enumerate(entries[:5])]
        bottom_slice = entries[-5:] if total > 5 else []
        bottom_start = total - len(bottom_slice)
        bottom = [_entry(bottom_start + i + 1, e) for i, e in enumerate(bottom_slice)]

        boards.append(
            {
                "key": kpi.key, "label": kpi.label, "group": kpi.group, "type": "ratio",
                "top": top, "bottom": bottom,
                "my_rank": (my_index + 1) if my_index is not None else None,
                "total": total,
            }
        )
    return boards


@router.get("/leaderboards")
def get_vswt_leaderboards(
    week: Optional[int] = Query(None),
    group: Optional[str] = Query(None),
    mode: Literal["latest", "consistency"] = Query(
        "latest", description="'latest' = this week's values; 'consistency' = average rank across every week on file."
    ),
    auth: AuthContext = Depends(get_auth_context),
    session: Session = Depends(get_session),
):
    # Leaderboards don't require the logged-in shop to have a shop_number — matches the
    # reference app, where leaderboards are visible even before "my shop" is known.
    shop_number = _shop_number_for(auth, session)
    weeks = _all_weeks(session)
    if not weeks:
        return {"available": False, "reason": "no_data"}
    kpis = KPI_DEFS if not group or group == "All" else [k for k in KPI_DEFS if k.group == group]

    if mode == "consistency":
        return {
            "available": True,
            "mode": "consistency",
            "week": weeks[-1],
            "weeks": weeks,
            "groups": ["All"] + list(KPI_GROUPS),
            "boards": _consistency_boards(session, weeks, kpis, shop_number),
        }

    target_week = week if week in weeks else weeks[-1]
    week_rows = _week_rows(session, target_week)

    boards = []
    for kpi in kpis:
        ranked = sorted(
            (r for r in week_rows if getattr(r, kpi.key) is not None),
            key=lambda r: getattr(r, kpi.key),
            reverse=True,
        )
        top = ranked[:5]
        bottom = ranked[-5:] if len(ranked) > 5 else []
        bottom_start = len(ranked) - len(bottom)
        my_index = (
            next((i for i, r in enumerate(ranked) if r.shop_number == shop_number), None)
            if shop_number
            else None
        )
        boards.append(
            {
                "key": kpi.key,
                "label": kpi.label,
                "group": kpi.group,
                "type": kpi.type,
                "top": [
                    {
                        "rank": i + 1, "shop_number": r.shop_number, "shop_name": r.shop_name,
                        "value": getattr(r, kpi.key), "is_me": r.shop_number == shop_number,
                    }
                    for i, r in enumerate(top)
                ],
                "bottom": [
                    {
                        "rank": bottom_start + i + 1,
                        "shop_number": r.shop_number,
                        "shop_name": r.shop_name,
                        "value": getattr(r, kpi.key), "is_me": r.shop_number == shop_number,
                    }
                    for i, r in enumerate(bottom)
                ],
                "my_rank": (my_index + 1) if my_index is not None else None,
                "total": len(ranked),
            }
        )

    return {
        "available": True,
        "mode": "latest",
        "week": target_week,
        "weeks": weeks,
        "groups": ["All"] + list(KPI_GROUPS),
        "boards": boards,
    }


@router.get("/trends")
def get_vswt_trends(
    weeks_back: int = Query(8, ge=1, le=104),
    shop_number: Optional[str] = Query(
        None, description="View another Minit shop's trends instead of your own (you must be a Minit shop yourself)."
    ),
    auth: AuthContext = Depends(get_auth_context),
    session: Session = Depends(get_session),
):
    my_shop_number = _shop_number_for(auth, session)
    if my_shop_number is None:
        return {"available": False, "reason": "no_shop_number"}
    target_shop_number = shop_number or my_shop_number
    all_weeks = _all_weeks(session)
    if not all_weeks:
        return {"available": False, "reason": "no_data"}
    weeks = all_weeks[-weeks_back:]

    sales_series = []
    rank_series = []
    found_any = False
    target_name = None
    target_area = None
    for w in weeks:
        week_rows = _week_rows(session, w)
        target_row = _find_shop(week_rows, target_shop_number)
        peers = _peer_rows(week_rows)
        if target_row is not None:
            found_any = True
            target_name = target_row.shop_name
            target_area = target_row.area_name
        sales_series.append(
            {
                "week": w,
                "shop": target_row.sales_ty if target_row else None,
                "region_avg": _average([r.sales_ty for r in week_rows]),
                "peer_avg": _average([r.sales_ty for r in peers]),
            }
        )
        rank_series.append(
            {"week": w, "rank": _rank_of(week_rows, "sales_ty", target_row.sales_ty) if target_row else None}
        )

    if not found_any:
        return {"available": False, "reason": "shop_not_found"}

    latest = weeks[-1]
    latest_rows = _week_rows(session, latest)
    latest_target_row = _find_shop(latest_rows, target_shop_number)
    category_series = [
        {
            "name": label,
            "shop": getattr(latest_target_row, key) if latest_target_row else None,
            "region_avg": _average([getattr(r, key) for r in latest_rows]),
        }
        for key, label in CATEGORY_SALES_KEYS
    ]

    return {
        "available": True,
        "shop_number": target_shop_number,
        "shop_name": target_name,
        "area_name": target_area,
        "viewing_own_shop": target_shop_number == my_shop_number,
        "weeks": weeks,
        "latest_week": latest,
        "sales_series": sales_series,
        "rank_series": rank_series,
        "category_series": category_series,
        "region_size": len(latest_rows),
    }


# ── Weekly report builder ────────────────────────────────────────────────────────────────
# Lets a shop hand-pick a handful of other shops (e.g. their own franchisee group) and get one
# week's numbers for just those shops laid out together — for pasting into a group chat, not for
# browsing the whole region like the Directory does.

def _parse_shop_numbers(shop_numbers: str) -> list[str]:
    # De-dupe while preserving the order the caller picked them in, so the report reads the same
    # order the user built it in rather than region sort order.
    seen: set[str] = set()
    out: list[str] = []
    for sn in shop_numbers.split(","):
        sn = sn.strip()
        if sn and sn not in seen:
            seen.add(sn)
            out.append(sn)
    return out


def _weekly_report_data(
    week_rows: list[VswtWeeklyShopMetric],
    shop_numbers: list[str],
    my_shop_number: Optional[str],
    compare_within_selection: bool = False,
) -> dict[str, Any]:
    """Shared by the JSON preview and the PDF export below, so both always show the same numbers.
    Comprehensive by design: every KPI HQ tracks gets a value *and* a rank for every selected
    shop, not just the Headline group — the report is meant to stand on its own without anyone
    needing to flip back to Rankings for the rest of the picture.

    `compare_within_selection` swaps what ranks are computed against: normally (False) every rank
    is the shop's position in the *whole region* — same numbers as the rest of the VSWT tabs. Set
    True to rank shops only against each other (e.g. "who's top of our own franchisee group this
    week"), which needs the selected shops resolved first so the ranking pool is just them."""
    by_number = {r.shop_number: r for r in week_rows}
    selected_rows: list[VswtWeeklyShopMetric] = []
    missing: list[str] = []
    for sn in shop_numbers:
        r = by_number.get(sn)
        if r is None:
            missing.append(sn)
        else:
            selected_rows.append(r)

    rank_pool = selected_rows if compare_within_selection else week_rows

    shops: list[dict[str, Any]] = []
    for r in selected_rows:
        values: dict[str, Optional[float]] = {}
        ranks: dict[str, Optional[int]] = {}
        for kpi in KPI_DEFS:
            v = getattr(r, kpi.key)
            values[kpi.key] = v
            ranks[kpi.key] = _rank_of(rank_pool, kpi.key, v)
        ranked = [rk for rk in ranks.values() if rk is not None]
        shops.append(
            {
                "shop_number": r.shop_number,
                "shop_name": r.shop_name,
                "area_name": r.area_name,
                "is_me": r.shop_number == my_shop_number,
                "sales_value": r.sales_ty,
                "sales_rank": ranks.get("sales_ty"),
                "customer_value": r.customer_ty,
                "jobs_value": r.jobs_ty,
                # Average rank across every tracked KPI — a single composite "how's this shop
                # doing overall" number alongside the headline sales rank.
                "overall_avg_rank": (sum(ranked) / len(ranked)) if ranked else None,
                "values": values,
                "ranks": ranks,
            }
        )
    # Best sales first — reads like a mini leaderboard for the group, nulls sink to the bottom.
    shops.sort(key=lambda s: (s["sales_value"] is None, -(s["sales_value"] or 0)))

    sales_vals = [s["sales_value"] for s in shops if s["sales_value"] is not None]
    customer_vals = [s["customer_value"] for s in shops if s["customer_value"] is not None]
    jobs_vals = [s["jobs_value"] for s in shops if s["jobs_value"] is not None]
    sales_ranks = [s["sales_rank"] for s in shops if s["sales_rank"] is not None]
    totals = {
        "sales": sum(sales_vals) if sales_vals else None,
        "customers": sum(customer_vals) if customer_vals else None,
        "jobs": sum(jobs_vals) if jobs_vals else None,
        "avg_sales_rank": (sum(sales_ranks) / len(sales_ranks)) if sales_ranks else None,
    }
    return {
        "shops": shops,
        "missing_shop_numbers": missing,
        "totals": totals,
        "rank_pool_size": len(rank_pool),
    }


@router.get("/weekly-report")
def get_vswt_weekly_report(
    week: Optional[int] = Query(None),
    shop_numbers: str = Query(..., description="Comma-separated shop numbers to include, in the order picked."),
    compare_within_selection: bool = Query(
        False, description="Rank shops only against each other instead of the whole region."
    ),
    auth: AuthContext = Depends(get_auth_context),
    session: Session = Depends(get_session),
):
    my_shop_number = _shop_number_for(auth, session)
    if my_shop_number is None:
        return {"available": False, "reason": "no_shop_number"}
    weeks = _all_weeks(session)
    if not weeks:
        return {"available": False, "reason": "no_data"}
    target_week = week if week in weeks else weeks[-1]

    numbers = _parse_shop_numbers(shop_numbers)
    if not numbers:
        raise HTTPException(status_code=400, detail="Pick at least one shop for the report.")

    week_rows = _week_rows(session, target_week)
    data = _weekly_report_data(week_rows, numbers, my_shop_number, compare_within_selection)
    if not data["shops"]:
        return {"available": False, "reason": "shop_not_found", "week": target_week}

    return {
        "available": True,
        "week": target_week,
        "weeks": weeks,
        "region_size": len(week_rows),
        "compare_within_selection": compare_within_selection,
        "groups": KPI_GROUPS,
        "kpis": [{"key": k.key, "label": k.label, "group": k.group, "type": k.type} for k in KPI_DEFS],
        **data,
    }


@router.get("/weekly-report/pdf")
def get_vswt_weekly_report_pdf(
    week: Optional[int] = Query(None),
    shop_numbers: str = Query(..., description="Comma-separated shop numbers to include, in the order picked."),
    title: str = Query("Weekly Regional Report", description="Report title, e.g. your franchisee group's name."),
    compare_within_selection: bool = Query(
        False, description="Rank shops only against each other instead of the whole region."
    ),
    auth: AuthContext = Depends(get_auth_context),
    session: Session = Depends(get_session),
):
    my_shop_number = _shop_number_for(auth, session)
    if my_shop_number is None:
        raise HTTPException(status_code=404, detail="This shop isn't linked to a VSWT shop number yet.")
    weeks = _all_weeks(session)
    if not weeks:
        raise HTTPException(status_code=404, detail="No regional data has been uploaded yet.")
    target_week = week if week in weeks else weeks[-1]

    numbers = _parse_shop_numbers(shop_numbers)
    if not numbers:
        raise HTTPException(status_code=400, detail="Pick at least one shop for the report.")

    week_rows = _week_rows(session, target_week)
    data = _weekly_report_data(week_rows, numbers, my_shop_number, compare_within_selection)
    if not data["shops"]:
        raise HTTPException(status_code=404, detail="None of the selected shops were found in this week's data.")

    pdf_bytes = build_weekly_report_pdf(
        title=title.strip() or "Weekly Regional Report",
        week=target_week,
        region_size=len(week_rows),
        compare_within_selection=compare_within_selection,
        rank_pool_size=data["rank_pool_size"],
        groups=KPI_GROUPS,
        kpis=KPI_DEFS,
        shops=data["shops"],
        totals=data["totals"],
        generated_on=date.today(),
    )
    filename = f"weekly-report-week-{target_week}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
