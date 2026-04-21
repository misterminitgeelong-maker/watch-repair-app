"""
CSV import endpoint – accepts an uploaded CSV file and imports
rows into watch repairs, shoe repairs, or mobile services (auto key), based on import_target.
"""

import csv
import io
import logging
import re
from datetime import date, datetime, timezone
from pathlib import Path
from uuid import UUID, uuid4

logger = logging.getLogger(__name__)

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Request, Query
from sqlalchemy import delete as sa_delete, func
from sqlmodel import Session, select
import openpyxl
import xlrd

from ..database import get_session
from ..dependencies import AuthContext, PLAN_FEATURES, enforce_plan_limit, get_auth_context
from ..config import settings
from ..limiter import limiter
from ..models import (
    Approval,
    Attachment,
    AutoKeyInvoice,
    AutoKeyJob,
    AutoKeyQuote,
    AutoKeyQuoteLineItem,
    Customer,
    CustomerAccountMembership,
    ImportLog,
    ImportLogDetail,
    ImportSummaryResponse,
    Invoice,
    JobStatusHistory,
    Payment,
    Quote,
    QuoteLineItem,
    RepairJob,
    SmsLog,
    Watch,
    WorkLog,
    Shoe,
    ShoeRepairJob,
    ShoeRepairJobItem,
    ShoeRepairJobShoe,
    ShoeJobStatusHistory,
    Tenant,
)

router = APIRouter(prefix="/v1/import", tags=["import"])


def get_import_csv_rate_limit() -> str:
    return settings.rate_limit_import_csv


# ── Helpers (ported from import_csv.py) ────────────────────────────────────────

def _parse_date(raw: str) -> date | None:
    if not raw or not raw.strip():
        return None
    raw = raw.strip()

    # Excel date serial (e.g. 44927 or 44927.0)
    try:
        n = float(raw)
        if 1000 < n < 100000 and n == int(n):
            from datetime import timedelta
            d = date(1899, 12, 30) + timedelta(days=int(n))
            return d
    except (ValueError, TypeError):
        pass

    for pattern, groups in [
        (r"^(\d{1,2})/(\d{1,2})/(\d{2})$", "dmy2"),
        (r"^(\d{1,2})/(\d{1,2})/(\d{4})$", "dmy4"),
        (r"^(\d{4})-(\d{1,2})-(\d{1,2})$", "ymd"),
        (r"^(\d{1,2})\.(\d{1,2})\.(\d{2,4})$", "dmy_dot"),
        (r"^(\d{1,2})/+(\d{1,2})/+(\d{2,4})$", "dmy_slash"),
    ]:
        m = re.match(pattern, raw)
        if not m:
            continue
        if groups == "ymd":
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        else:
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        if 1 <= mo <= 12 and 1 <= d <= 31:
            try:
                return date(y, mo, d)
            except ValueError:
                continue
    return None


def _parse_date_flexible(raw: str) -> date | None:
    """Like ``_parse_date`` but also tries the first whitespace-separated token (e.g. ``23.7.25 9.30am``)."""
    d = _parse_date(raw or "")
    if d:
        return d
    head = (raw or "").strip().split()[0] if (raw or "").strip() else ""
    return _parse_date(head) if head else None


def _normalize_phone(raw: str) -> str | None:
    if not raw or not raw.strip():
        return None
    digits = re.sub(r"\D", "", raw.strip())
    if len(digits) > 12:
        digits = digits[:10]
    if len(digits) == 9 and digits[0] in ("4", "3"):
        digits = "0" + digits
    if len(digits) < 8:
        return None
    return digits


# PostgreSQL INTEGER max; clamp to avoid overflow from corrupted Excel values
_MAX_CENTS = 2_147_483_647


def _dollars_to_cents(raw: str) -> int:
    if not raw or not raw.strip():
        return 0
    s = raw.strip().lower()
    if any(token in s for token in ["n/c", "nc", "quote"]):
        return 0

    # Handles values like "2x69.95" or "2 x $49.95 - 10%"
    m = re.search(r"(\d+)\s*x\s*\$?\s*(\d+(?:\.\d+)?)", s)
    if m:
        qty = int(m.group(1))
        unit = float(m.group(2))
        total = qty * unit
        discount = re.search(r"-\s*(\d+(?:\.\d+)?)\s*%", s)
        if discount:
            total *= 1 - (float(discount.group(1)) / 100)
        return min(int(round(total * 100)), _MAX_CENTS)

    numbers = [float(n) for n in re.findall(r"\d+(?:\.\d+)?", s)]
    if not numbers:
        return 0
    return min(int(round(numbers[0] * 100)), _MAX_CENTS)


def _clean_name(raw: str) -> str | None:
    if not raw or not raw.strip():
        return None
    name = re.sub(r"\s+", " ", raw.strip())
    lower_name = name.lower()

    # Common placeholders/comments that are not valid customer names.
    invalid_name_fragments = [
        "already collected",
        "has been collected",
        "picked up",
        "cancelled",
        "unknown",
    ]
    if any(fragment in lower_name for fragment in invalid_name_fragments):
        return None

    if re.match(r"^\d+$", name):
        return None
    if re.match(r"^\d{1,2}/\d{1,2}/\d{2,4}", name):
        return None
    if len(name) < 2:
        return None
    return name


def _get_first(row: dict[str, str], keys: list[str]) -> str:
    """Return the first non-empty value found in the provided column keys."""
    for key in keys:
        value = row.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _normalize_key(key: str) -> str:
    k = (key or "").strip().lower()
    k = re.sub(r"[^a-z0-9]+", "_", k)
    return k.strip("_")


def _normalize_row_keys(row: dict[str, str]) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for key, value in row.items():
        normalized[_normalize_key(key)] = (value or "").strip()

    # First column with no header (Excel: _col0→col0, CSV: "") often contains ticket/job numbers.
    for blank_key in ("", "_col0", "col0"):
        val = (row.get(blank_key) or "").strip()
        if val and "original_job_id" not in normalized:
            try:
                if re.match(r"^[\d.]+$", val):
                    normalized["original_job_id"] = val
                    break
            except Exception:
                pass
    return normalized


def _to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _load_csv_rows(raw_bytes: bytes) -> list[dict[str, str]]:
    text = ""
    for encoding in ("utf-8-sig", "utf-16", "latin-1"):
        try:
            text = raw_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue

    if not text:
        raise HTTPException(status_code=400, detail="Unable to decode CSV file. Please export as UTF-8 CSV.")

    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    except csv.Error:
        reader = csv.DictReader(io.StringIO(text))

    rows = list(reader)
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV header row could not be read. Ensure the first row contains column names.")
    return rows


def _score_repair_header_row(header_row: tuple) -> int:
    """Prefer worksheets that look like a repair/job log (flexible column names)."""
    keys = {_normalize_key(_to_text(v) or f"_col{i}") for i, v in enumerate(header_row)}
    score = 0
    if {"name", "customer", "client_name", "customer_name"} & keys:
        score += 2
    if {"number", "phone", "phone_number", "mobile"} & keys:
        score += 2
    if {"brand", "brand_case_numbers", "make_model", "model"} & keys:
        score += 1
    if {"date", "date_in", "date_time", "datetime", "created_at", "date_recieved", "date_received"} & keys:
        score += 1
    if "quote" in keys or "quote_price" in keys:
        score += 1
    if {"repair_notes", "notes", "notes_on_job", "description", "job_description", "other_notes"} & keys:
        score += 1
    if {"status", "ga_ng_collected", "job_status"} & keys:
        score += 1
    if {"address", "job_address"} & keys:
        score += 1
    if {"docket_number", "ticket", "ticket_number", "col0", "_col0"} & keys or any(
        k in keys for k in keys if k.startswith("col") and k[3:].isdigit()
    ):
        score += 1
    return score


def _pick_xlsx_sheet(workbook, sheet_name: str | None):
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            names = ", ".join(workbook.sheetnames[:12])
            more = f" (+{len(workbook.sheetnames) - 12} more)" if len(workbook.sheetnames) > 12 else ""
            raise HTTPException(
                status_code=400,
                detail=f'Worksheet "{sheet_name}" not found. Available: {names}{more}',
            )
        return workbook[sheet_name]
    best_ws = None
    best_score = -1
    for ws in workbook.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        sc = _score_repair_header_row(rows[0])
        if sc > best_score:
            best_score = sc
            best_ws = ws
    if best_ws is not None and best_score >= 3:
        return best_ws
    return workbook.active


def _load_xlsx_rows(raw_bytes: bytes, sheet_name: str | None = None) -> tuple[list[dict[str, str]], str]:
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
        sheet = _pick_xlsx_sheet(workbook, sheet_name)
        if sheet is None:
            raise HTTPException(status_code=400, detail="Excel file has no worksheets.")
        all_rows = list(sheet.iter_rows(values_only=True))
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to read uploaded Excel file")
        raise HTTPException(
            status_code=400,
            detail="Could not read the Excel file. Try saving as CSV (UTF-8) or a fresh .xlsx.",
        ) from exc

    if not all_rows:
        return [], sheet.title

    header_row = all_rows[0]
    header = [_to_text(v) or f"_col{i}" for i, v in enumerate(header_row)]
    # Ensure unique header keys (Excel often has blank columns)
    seen: set[str] = set()
    for i, h in enumerate(header):
        if h in seen:
            header[i] = f"{h}_{i}"
        seen.add(header[i])

    result: list[dict[str, str]] = []
    for row in all_rows[1:]:
        if not any(_to_text(c) for c in row):
            continue
        row_len = len(row)
        hdr_len = len(header)
        cells = [_to_text(row[idx]) if idx < row_len else "" for idx in range(hdr_len)]
        result.append({header[i]: cells[i] for i in range(hdr_len)})
    return result, sheet.title


def _load_xls_rows(raw_bytes: bytes) -> list[dict[str, str]]:
    workbook = xlrd.open_workbook(file_contents=raw_bytes)
    sheet = workbook.sheet_by_index(0)
    if sheet.nrows == 0:
        return []

    header = [_to_text(sheet.cell_value(0, col)) for col in range(sheet.ncols)]
    output: list[dict[str, str]] = []
    for row_idx in range(1, sheet.nrows):
        row_values = [_to_text(sheet.cell_value(row_idx, col)) for col in range(sheet.ncols)]
        if not any(row_values):
            continue
        output.append({header[idx]: value for idx, value in enumerate(row_values)})
    return output


def _load_rows(
    filename: str, raw_bytes: bytes, sheet_name: str | None = None
) -> tuple[list[dict[str, str]], str, str | None]:
    ext = Path(filename).suffix.lower()
    if ext == ".csv":
        return _load_csv_rows(raw_bytes), "csv", None
    if ext in {".xlsx", ".xlsm"}:
        rows, title = _load_xlsx_rows(raw_bytes, sheet_name)
        return rows, "xlsx", title
    if ext == ".xls":
        return _load_xls_rows(raw_bytes), "xls", None

    # Be tolerant of renamed/missing extensions by sniffing common formats.
    if raw_bytes.startswith(b"PK"):
        rows, title = _load_xlsx_rows(raw_bytes, sheet_name)
        return rows, "xlsx", title
    if raw_bytes.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"):
        return _load_xls_rows(raw_bytes), "xls", None

    try:
        return _load_csv_rows(raw_bytes), "csv", None
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail="Only .csv, .xlsx, .xls, and .xlsm files are accepted. If this is CSV, re-save as UTF-8 CSV.",
        ) from exc


_STATUS_MAP = {
    "collected": "collected",
    "done": "completed",
    "ga": "completed",  # Good as New
    "ng": "no_go",     # No Good
    "in_repair": "working_on",
    "in repair": "working_on",
    "working on": "working_on",
    "ready": "awaiting_collection",
    "awaiting collection": "awaiting_collection",
    "intake": "awaiting_go_ahead",
    "awaiting go ahead": "awaiting_go_ahead",
    "go ahead": "go_ahead",
    "no go": "no_go",
    "completed": "completed",
    "awaiting parts": "awaiting_parts",
    "parts to be ordered": "parts_to_order",
    "sent to labanda": "sent_to_labanda",
    "service": "service",
    "cancelled": "no_go",
    "approved": "go_ahead",
}

# Status cells like "Sent Back Completed" → slug sent_back_completed
_SLUG_STATUS_ALIASES: dict[str, str] = {
    "sent_back_completed": "completed",
    "ready_for_collection": "awaiting_collection",
}


def _status_slug(status_raw: str) -> str:
    s = (status_raw or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return re.sub(r"_+", "_", s).strip("_")


def _allocate_import_job_number(usage: dict[str, int], ticket_stem: str | None, job_seq: int) -> str:
    """Assign IMP-{ticket} or IMP-{seq}; duplicate tickets become IMP-{ticket}-2, -3, …"""
    if ticket_stem:
        base = f"IMP-{ticket_stem}"
    else:
        base = f"IMP-{job_seq:05d}"
    n = usage.get(base, 0)
    usage[base] = n + 1
    if n == 0:
        return base
    return f"{base}-{n + 1}"


def _reserve_import_job_number(preferred: str, used_job_numbers: set[str]) -> str:
    """
    Pick a job_number not in ``used_job_numbers``, starting from preferred (in-file slot).
    Updates ``used_job_numbers`` with the chosen value (in-memory; avoids per-row DB round-trips).
    """
    m = re.match(r"^(IMP-\d+)(?:-(\d+))?$", preferred)
    if m:
        root = m.group(1)
        k = int(m.group(2)) if m.group(2) else 1
    else:
        root = preferred
        k = 1
    while True:
        candidate = root if k == 1 else f"{root}-{k}"
        if candidate not in used_job_numbers:
            used_job_numbers.add(candidate)
            return candidate
        k += 1


def _infer_job_status(status_raw: str, notes_raw: str) -> str:
    status = (status_raw or "").strip().lower()
    notes = (notes_raw or "").strip().lower()
    slug = _status_slug(status_raw)

    if slug in _SLUG_STATUS_ALIASES:
        return _SLUG_STATUS_ALIASES[slug]
    if slug in _STATUS_MAP:
        return _STATUS_MAP[slug]
    if status in _STATUS_MAP:
        return _STATUS_MAP[status]

    if status == "delivered":
        if any(token in notes for token in ["collected", "picked up", "has been collected", "watch collect"]):
            return "collected"
        if any(token in notes for token in ["awaiting collection", "ready for collection", "ready in drawer", "messaged for collection"]):
            return "awaiting_collection"
        return "completed"

    return "awaiting_go_ahead"


def _infer_auto_key_status(status_raw: str, notes_raw: str) -> str:
    """Map spreadsheet status to AutoKeyJob.status (JobStatus literals)."""
    slug = _status_slug(status_raw)
    s = (status_raw or "").strip().lower()

    mobile_slug_map: dict[str, str] = {
        "booked": "booked",
        "confirmed": "booked",
        "scheduled": "booked",
        "pending_booking": "pending_booking",
        "awaiting_booking": "pending_booking",
        "en_route": "en_route",
        "on_the_way": "en_route",
        "driving": "en_route",
        "on_site": "on_site",
        "arrived": "on_site",
        "at_site": "on_site",
        "awaiting_customer_details": "awaiting_customer_details",
    }
    if slug in mobile_slug_map:
        return mobile_slug_map[slug]
    if s in mobile_slug_map:
        return mobile_slug_map[s]

    base = _infer_job_status(status_raw, notes_raw)
    if base == "awaiting_go_ahead" and not (status_raw or "").strip():
        return "awaiting_quote"
    if base in {
        "awaiting_go_ahead",
        "go_ahead",
        "working_on",
        "awaiting_parts",
        "parts_to_order",
        "sent_to_labanda",
        "quoted_by_labanda",
        "service",
        "completed",
        "awaiting_collection",
        "collected",
        "no_go",
        "en_route",
        "on_site",
        "pending_booking",
        "booked",
        "awaiting_customer_details",
    }:
        return base
    return "awaiting_quote"


def _effective_clear_tabs(import_target: str, clear_tabs: list[str], replace_existing: bool) -> list[str]:
    """When replacing, only clear the module being imported unless the caller passes a subset."""
    allowed = {"watch": "watch", "shoe": "shoe", "mobile": "auto_key"}[import_target]
    filtered = [t for t in clear_tabs if t == allowed]
    if replace_existing and not filtered:
        return [allowed]
    return filtered


def _auth_has_import_target(auth: AuthContext, import_target: str) -> bool:
    if auth.role == "platform_admin":
        return True
    features = PLAN_FEATURES.get(auth.plan_code, PLAN_FEATURES["pro"])
    if import_target == "watch":
        return "watch" in features
    if import_target == "shoe":
        return "shoe" in features
    if import_target == "mobile":
        return "auto_key" in features
    return False


def _infer_customer_account_id(session: Session, tenant_id: UUID, customer_id: UUID) -> UUID | None:
    row = session.exec(
        select(CustomerAccountMembership)
        .where(CustomerAccountMembership.tenant_id == tenant_id)
        .where(CustomerAccountMembership.customer_id == customer_id)
        .order_by(CustomerAccountMembership.created_at)
    ).first()
    return row.customer_account_id if row else None


def _clear_tenant_importable_data(session: Session, tenant_id, clear_tabs: list[str], *, delete_customers: bool = True) -> None:
    """Delete tenant-scoped operational data before replacement import, for selected tabs.

    When clearing only ``watch``, set ``delete_customers=False`` so customers shared with
    shoe or mobile services jobs are not removed (avoids broken FKs and matches tab-scoped imports).
    """
    import_log_ids = session.exec(select(ImportLog.id).where(ImportLog.tenant_id == tenant_id)).all()
    if import_log_ids:
        session.exec(sa_delete(ImportLogDetail).where(ImportLogDetail.import_log_id.in_(import_log_ids)))
        session.exec(sa_delete(ImportLog).where(ImportLog.id.in_(import_log_ids)))

    # --- SHOE REPAIR TABLES ---
    if "shoe" in clear_tabs:
        session.exec(
            sa_delete(Attachment).where(Attachment.tenant_id == tenant_id).where(Attachment.shoe_repair_job_id.isnot(None))
        )
        session.exec(
            sa_delete(SmsLog).where(SmsLog.tenant_id == tenant_id).where(SmsLog.shoe_repair_job_id.isnot(None))
        )
        session.exec(sa_delete(ShoeJobStatusHistory).where(ShoeJobStatusHistory.tenant_id == tenant_id))
        session.exec(sa_delete(ShoeRepairJobItem).where(ShoeRepairJobItem.tenant_id == tenant_id))
        session.exec(sa_delete(ShoeRepairJobShoe).where(ShoeRepairJobShoe.tenant_id == tenant_id))
        session.exec(sa_delete(ShoeRepairJob).where(ShoeRepairJob.tenant_id == tenant_id))
        session.exec(sa_delete(Shoe).where(Shoe.tenant_id == tenant_id))

    # --- AUTO KEY TABLES (before watch clear deletes Customer rows) ---
    if "auto_key" in clear_tabs:
        session.exec(
            sa_delete(Attachment).where(Attachment.tenant_id == tenant_id).where(Attachment.auto_key_job_id.isnot(None))
        )
        session.exec(sa_delete(AutoKeyInvoice).where(AutoKeyInvoice.tenant_id == tenant_id))
        session.exec(sa_delete(AutoKeyQuoteLineItem).where(AutoKeyQuoteLineItem.tenant_id == tenant_id))
        session.exec(sa_delete(AutoKeyQuote).where(AutoKeyQuote.tenant_id == tenant_id))
        session.exec(sa_delete(AutoKeyJob).where(AutoKeyJob.tenant_id == tenant_id))

    # --- WATCH REPAIR TABLES ---
    if "watch" in clear_tabs:
        session.exec(sa_delete(Attachment).where(Attachment.tenant_id == tenant_id))
        session.exec(sa_delete(WorkLog).where(WorkLog.tenant_id == tenant_id))
        session.exec(sa_delete(JobStatusHistory).where(JobStatusHistory.tenant_id == tenant_id))
        session.exec(sa_delete(SmsLog).where(SmsLog.tenant_id == tenant_id))
        session.exec(sa_delete(QuoteLineItem).where(QuoteLineItem.tenant_id == tenant_id))
        session.exec(sa_delete(Approval).where(Approval.tenant_id == tenant_id))
        session.exec(sa_delete(Payment).where(Payment.tenant_id == tenant_id))
        session.exec(sa_delete(Invoice).where(Invoice.tenant_id == tenant_id))
        session.exec(sa_delete(Quote).where(Quote.tenant_id == tenant_id))
        session.exec(sa_delete(RepairJob).where(RepairJob.tenant_id == tenant_id))
        session.exec(sa_delete(Watch).where(Watch.tenant_id == tenant_id))
        if delete_customers:
            session.exec(
                sa_delete(CustomerAccountMembership).where(CustomerAccountMembership.tenant_id == tenant_id)
            )
            session.exec(sa_delete(Customer).where(Customer.tenant_id == tenant_id))


# ── Endpoint ───────────────────────────────────────────────────────────────────


@router.post("/csv", response_model=ImportSummaryResponse)
@limiter.limit(get_import_csv_rate_limit)
async def import_csv(
    request: Request,
    file: UploadFile = File(...),
    replace_existing: bool = False,
    dry_run: bool = False,
    sheet_name: str | None = Query(
        None,
        description='Excel worksheet name (optional). If omitted, the importer picks the sheet that best matches a repair log.',
    ),
    clear_tabs: list[str] = Query(["watch"], description="Tab(s) to clear before import when replace_existing is true."),
    import_target: str = Query(
        "watch",
        description="Import destination: watch (watch repairs), shoe (shoe repair jobs), mobile (mobile services / auto key jobs).",
        pattern="^(watch|shoe|mobile)$",
    ),
    auth: AuthContext = Depends(get_auth_context),
    session: Session = Depends(get_session),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="File name is required")

    if import_target not in ("watch", "shoe", "mobile"):
        raise HTTPException(status_code=400, detail="import_target must be watch, shoe, or mobile")
    if not _auth_has_import_target(auth, import_target):
        raise HTTPException(
            status_code=403,
            detail=f"Your plan does not include imports for '{import_target}'.",
        )

    raw_bytes = await file.read()
    rows, file_type, source_sheet = _load_rows(file.filename, raw_bytes, sheet_name)
    if not rows:
        raise HTTPException(status_code=400, detail="Import file is empty.")

    tenant_id = auth.tenant_id
    tenant = session.get(Tenant, tenant_id)
    tenant_currency = (tenant.default_currency if tenant and tenant.default_currency else "AUD").upper()
    effective_clear = _effective_clear_tabs(import_target, clear_tabs, replace_existing)

    has_non_watch_customer_refs = bool(
        session.exec(select(AutoKeyJob.id).where(AutoKeyJob.tenant_id == tenant_id).limit(1)).first()
        or session.exec(select(Shoe.id).where(Shoe.tenant_id == tenant_id).limit(1)).first()
    )
    delete_customers_with_watch_clear = not has_non_watch_customer_refs

    if replace_existing and not dry_run:
        try:
            _clear_tenant_importable_data(
                session,
                tenant_id,
                effective_clear,
                delete_customers=delete_customers_with_watch_clear,
            )
            session.commit()
        except Exception as exc:
            session.rollback()
            logger.exception(
                "Could not clear existing data before import for tenant %s", auth.tenant_id
            )
            raise HTTPException(
                status_code=400,
                detail="Could not clear existing data before import. Please try again.",
            ) from exc

    import_log = None
    if not dry_run:
        import_log = ImportLog(
            tenant_id=tenant_id,
            uploaded_by_user_id=auth.user_id,
            file_name=file.filename,
            file_type=file_type,
            total_rows=len(rows),
        )
        session.add(import_log)
        session.commit()
        session.refresh(import_log)

    customer_cache: dict[tuple[str, str | None], Customer] = {}
    imported = 0
    skipped = 0
    job_seq = 0
    skipped_reasons: dict[str, int] = {}
    duplicate_customer_rows_in_file = 0
    job_number_usage: dict[str, int] = {}

    def log_skip(row_number: int, reason: str) -> None:
        nonlocal skipped
        skipped += 1
        skipped_reasons[reason] = skipped_reasons.get(reason, 0) + 1
        if import_log is not None:
            session.add(ImportLogDetail(import_log_id=import_log.id, row_number=row_number, skip_reason=reason))

    try:
        if import_target == "watch":
            existing_job_nums = session.exec(
                select(RepairJob.job_number).where(RepairJob.tenant_id == tenant_id)
            ).all()
            used_job_numbers: set[str] = {jn for jn in existing_job_nums if jn}

            # B-M2: count once before the loop, increment locally on each
            # successful create. Previously this ran a COUNT(*) per row.
            repair_job_running_count = int(
                session.exec(
                    select(func.count()).select_from(RepairJob).where(RepairJob.tenant_id == tenant_id)
                ).one()
            )

            for idx, row in enumerate(rows, start=1):
                row = _normalize_row_keys(row)
                original_job_id = _get_first(row, [
                    "col0", "_col0", "original_job_id", "job_id", "ticket", "ticket_number", "docket_number",
                    "docket", "job_number", "ticket_", "job_", "ref", "job_no", "ticket_no", "case_no", "no",
                ])
                team_member = _get_first(row, ["team_member", "salesperson", "staff", "assignee", "store", "source"])
                customer_name_raw = _get_first(row, ["customer_name", "customer", "client", "client_name", "name"])
                date_in_raw = _get_first(row, [
                    "date_in", "created_at", "created", "intake_date", "date", "date_received", "date_recieved",
                    "date_time", "datetime", "scheduled_at", "appointment",
                ])
                brand_case = _get_first(row, [
                    "brand_case_numbers", "brand", "watch_brand", "make_model", "model",
                    "to_do_on_job", "work_required",
                ])
                phone_raw = _get_first(row, ["phone_number", "phone", "mobile", "contact_phone", "number"])
                quote_raw = _get_first(row, ["quote_price", "quote", "estimate", "amount", "total"])
                cost_raw = _get_first(row, [
                    "cost_to_business", "our_cost", "cost", "job_cost", "internal_cost",
                ])
                status_raw = _get_first(row, [
                    "status", "job_status", "repair_status", "state", "ga_ng_collected",
                    "quote_sent_to_geelong",
                ])
                notes_raw = _get_first(row, [
                    "repair_notes", "notes_on_job", "notes", "description", "job_notes", "being_done",
                    "job_description", "other_notes",
                ])
                address_raw = _get_first(row, ["address", "customer_address", "street"])

                customer_name = _clean_name(customer_name_raw)
                if not customer_name:
                    customer_name = _clean_name(original_job_id)
                if not customer_name:
                    has_meaningful_data = any(
                        [brand_case, phone_raw, quote_raw, status_raw, notes_raw, original_job_id, address_raw]
                    )
                    if not has_meaningful_data:
                        log_skip(idx, "empty_row")
                        continue
                    if not brand_case and not phone_raw:
                        log_skip(idx, "missing_core_fields")
                        continue
                    customer_name = f"Unknown Customer {original_job_id or uuid4().hex[:8]}"

                phone = _normalize_phone(phone_raw)
                date_in = _parse_date_flexible(date_in_raw)
                status = _infer_job_status(status_raw, notes_raw)
                quote_cents = _dollars_to_cents(quote_raw)
                cost_cents = _dollars_to_cents(cost_raw)

                cache_key = (customer_name.lower(), phone)
                created_customer_id = None
                if cache_key in customer_cache:
                    customer = customer_cache[cache_key]
                    duplicate_customer_rows_in_file += 1
                else:
                    addr = (address_raw or "").strip() or None
                    customer = Customer(tenant_id=tenant_id, full_name=customer_name, phone=phone, address=addr)
                    session.add(customer)
                    session.flush()
                    customer_cache[cache_key] = customer
                    created_customer_id = customer.id

                watch = Watch(tenant_id=tenant_id, customer_id=customer.id, brand=brand_case or None)
                session.add(watch)
                session.flush()

                job_seq += 1
                ticket_stem: str | None = None
                if original_job_id:
                    ticket = original_job_id.strip()
                    if re.match(r"^[\d.]+$", ticket) and ticket.endswith(".0"):
                        ticket = ticket[:-2]
                    if ticket:
                        ticket_stem = ticket
                preferred_num = _allocate_import_job_number(job_number_usage, ticket_stem, job_seq)
                job_number = _reserve_import_job_number(preferred_num, used_job_numbers)

                title = f"Repair: {brand_case}" if brand_case else "Watch Repair"
                created_at = (
                    datetime(date_in.year, date_in.month, date_in.day, tzinfo=timezone.utc)
                    if date_in
                    else datetime.now(timezone.utc)
                )

                enforce_plan_limit(auth, "repair_job", repair_job_running_count)
                repair_job_running_count += 1

                job = RepairJob(
                    tenant_id=tenant_id,
                    watch_id=watch.id,
                    job_number=job_number,
                    title=title,
                    description=notes_raw or None,
                    priority="normal",
                    status=status,
                    salesperson=team_member or None,
                    deposit_cents=0,
                    pre_quote_cents=quote_cents or cost_cents,
                    cost_cents=cost_cents,
                    created_at=created_at,
                )
                session.add(job)
                session.flush()

                if quote_cents > 0:
                    quote_status = "approved" if status in {"completed", "awaiting_collection", "collected"} else "sent"
                    session.add(Quote(
                        tenant_id=tenant_id,
                        repair_job_id=job.id,
                        status=quote_status,
                        subtotal_cents=quote_cents,
                        tax_cents=0,
                        total_cents=quote_cents,
                        currency=tenant_currency,
                        created_at=created_at,
                    ))

                if import_log is not None:
                    session.add(
                        ImportLogDetail(
                            import_log_id=import_log.id,
                            row_number=idx,
                            created_repair_job_id=job.id,
                            created_customer_id=created_customer_id,
                        )
                    )

                imported += 1

        elif import_target == "shoe":
            existing_job_nums = session.exec(
                select(ShoeRepairJob.job_number).where(ShoeRepairJob.tenant_id == tenant_id)
            ).all()
            used_job_numbers = {jn for jn in existing_job_nums if jn}

            shoe_job_running_count = int(
                session.exec(
                    select(func.count()).select_from(ShoeRepairJob).where(ShoeRepairJob.tenant_id == tenant_id)
                ).one()
            )

            for idx, row in enumerate(rows, start=1):
                row = _normalize_row_keys(row)
                original_job_id = _get_first(row, [
                    "col0", "_col0", "original_job_id", "job_id", "ticket", "ticket_number", "docket_number",
                    "docket", "job_number", "ref", "job_no", "ticket_no",
                ])
                team_member = _get_first(row, ["team_member", "salesperson", "staff", "assignee", "store", "source"])
                customer_name_raw = _get_first(row, ["customer_name", "customer", "client", "client_name", "name"])
                date_in_raw = _get_first(row, [
                    "date_in", "created_at", "created", "intake_date", "date", "date_received", "date_recieved",
                    "date_time", "datetime",
                ])
                shoe_brand = _get_first(row, [
                    "shoe_brand", "brand", "make_model", "model", "style",
                ])
                shoe_type = _get_first(row, ["shoe_type", "type", "category"])
                color = _get_first(row, ["color", "colour"])
                phone_raw = _get_first(row, ["phone_number", "phone", "mobile", "contact_phone", "number"])
                quote_raw = _get_first(row, ["quote_price", "quote", "estimate", "amount", "total"])
                cost_raw = _get_first(row, ["cost_to_business", "our_cost", "cost", "job_cost", "internal_cost"])
                status_raw = _get_first(row, ["status", "job_status", "repair_status", "state"])
                notes_raw = _get_first(row, [
                    "repair_notes", "notes_on_job", "notes", "description", "job_notes", "being_done",
                    "job_description", "other_notes", "work_required", "to_do_on_job",
                ])
                title_raw = _get_first(row, ["title", "job_title", "summary"])
                address_raw = _get_first(row, ["address", "customer_address", "street"])

                customer_name = _clean_name(customer_name_raw) or _clean_name(original_job_id)
                if not customer_name:
                    has_data = any([shoe_brand, phone_raw, quote_raw, status_raw, notes_raw, original_job_id])
                    if not has_data:
                        log_skip(idx, "empty_row")
                        continue
                    if not shoe_brand and not phone_raw:
                        log_skip(idx, "missing_core_fields")
                        continue
                    customer_name = f"Unknown Customer {original_job_id or uuid4().hex[:8]}"

                phone = _normalize_phone(phone_raw)
                if not phone and not shoe_brand and not notes_raw:
                    log_skip(idx, "missing_core_fields")
                    continue

                cache_key = (customer_name.lower(), phone)
                created_customer_id = None
                if cache_key in customer_cache:
                    customer = customer_cache[cache_key]
                    duplicate_customer_rows_in_file += 1
                else:
                    addr = (address_raw or "").strip() or None
                    customer = Customer(tenant_id=tenant_id, full_name=customer_name, phone=phone, address=addr)
                    session.add(customer)
                    session.flush()
                    customer_cache[cache_key] = customer
                    created_customer_id = customer.id

                shoe = Shoe(
                    tenant_id=tenant_id,
                    customer_id=customer.id,
                    shoe_type=shoe_type or None,
                    brand=shoe_brand or None,
                    color=color or None,
                    description_notes=notes_raw or None,
                )
                session.add(shoe)
                session.flush()

                job_seq += 1
                ticket_stem = None
                if original_job_id:
                    ticket = original_job_id.strip()
                    if re.match(r"^[\d.]+$", ticket) and ticket.endswith(".0"):
                        ticket = ticket[:-2]
                    if ticket:
                        ticket_stem = f"SH{ticket}"
                preferred_num = _allocate_import_job_number(job_number_usage, ticket_stem, job_seq)
                job_number = _reserve_import_job_number(preferred_num, used_job_numbers)

                title = (title_raw.strip() if title_raw else "") or (
                    f"Shoe repair: {shoe_brand}" if shoe_brand else "Shoe repair"
                )
                date_in = _parse_date_flexible(date_in_raw)
                created_at = (
                    datetime(date_in.year, date_in.month, date_in.day, tzinfo=timezone.utc)
                    if date_in
                    else datetime.now(timezone.utc)
                )
                status = _infer_job_status(status_raw, notes_raw)
                quote_cents = _dollars_to_cents(quote_raw)
                cost_cents = _dollars_to_cents(cost_raw)
                job_cost = cost_cents or quote_cents
                qstat = "none"
                if quote_cents > 0:
                    qstat = "approved" if status in {"completed", "awaiting_collection", "collected", "go_ahead"} else "sent"

                ca_id = _infer_customer_account_id(session, tenant_id, customer.id)

                enforce_plan_limit(auth, "shoe_job", shoe_job_running_count)
                shoe_job_running_count += 1

                job = ShoeRepairJob(
                    tenant_id=tenant_id,
                    shoe_id=shoe.id,
                    customer_account_id=ca_id,
                    job_number=job_number,
                    title=title[:500],
                    description=notes_raw or None,
                    priority="normal",
                    status=status,
                    salesperson=team_member or None,
                    deposit_cents=0,
                    cost_cents=job_cost,
                    quote_status=qstat,
                    created_at=created_at,
                )
                session.add(job)
                session.flush()

                session.add(
                    ShoeJobStatusHistory(
                        tenant_id=tenant_id,
                        shoe_repair_job_id=job.id,
                        old_status=None,
                        new_status=status,
                        changed_by_user_id=auth.user_id,
                        change_note="Imported",
                    )
                )

                if import_log is not None:
                    session.add(
                        ImportLogDetail(
                            import_log_id=import_log.id,
                            row_number=idx,
                            created_customer_id=created_customer_id,
                        )
                    )
                imported += 1

        else:  # mobile (auto key)
            existing_job_nums = session.exec(
                select(AutoKeyJob.job_number).where(AutoKeyJob.tenant_id == tenant_id)
            ).all()
            used_job_numbers = {jn for jn in existing_job_nums if jn}

            auto_key_running_count = int(
                session.exec(
                    select(func.count()).select_from(AutoKeyJob).where(AutoKeyJob.tenant_id == tenant_id)
                ).one()
            )

            for idx, row in enumerate(rows, start=1):
                row = _normalize_row_keys(row)
                original_job_id = _get_first(row, [
                    "col0", "_col0", "original_job_id", "job_id", "ticket", "ticket_number", "docket_number",
                    "docket", "job_number", "ref", "job_no", "ticket_no",
                ])
                team_member = _get_first(row, ["team_member", "salesperson", "staff", "assignee", "store", "source"])
                customer_name_raw = _get_first(row, ["customer_name", "customer", "client", "client_name", "name"])
                date_in_raw = _get_first(row, [
                    "date_in", "created_at", "created", "intake_date", "date", "date_received", "date_recieved",
                    "date_time", "datetime", "scheduled_at", "appointment",
                ])
                phone_raw = _get_first(row, ["phone_number", "phone", "mobile", "contact_phone", "number"])
                quote_raw = _get_first(row, ["quote_price", "quote", "estimate", "amount", "total"])
                cost_raw = _get_first(row, ["cost_to_business", "our_cost", "cost", "job_cost", "internal_cost"])
                status_raw = _get_first(row, ["status", "job_status", "state"])
                notes_raw = _get_first(row, [
                    "repair_notes", "notes_on_job", "notes", "description", "job_notes",
                    "job_description", "other_notes", "tech_notes",
                ])
                address_raw = _get_first(row, ["address", "job_address", "street", "location"])
                vehicle_make = _get_first(row, ["vehicle_make", "make", "car_make"])
                vehicle_model = _get_first(row, ["vehicle_model", "model", "car_model"])
                vehicle_year_raw = _get_first(row, ["vehicle_year", "year"])
                rego = _get_first(row, ["registration", "registration_plate", "rego", "plate"])
                job_type = _get_first(row, ["job_type", "service", "service_type"])
                title_raw = _get_first(row, ["title", "job_title", "summary"])
                work_line = _get_first(row, ["to_do_on_job", "work_required", "job_description"])

                customer_name = _clean_name(customer_name_raw) or _clean_name(original_job_id)
                if not customer_name:
                    has_data = any(
                        [phone_raw, quote_raw, status_raw, notes_raw, address_raw, work_line, vehicle_make, original_job_id]
                    )
                    if not has_data:
                        log_skip(idx, "empty_row")
                        continue
                    if not phone_raw and not address_raw and not work_line:
                        log_skip(idx, "missing_core_fields")
                        continue
                    customer_name = f"Unknown Customer {original_job_id or uuid4().hex[:8]}"

                phone = _normalize_phone(phone_raw)
                if not phone and not address_raw and not work_line:
                    log_skip(idx, "missing_core_fields")
                    continue

                cache_key = (customer_name.lower(), phone)
                created_customer_id = None
                if cache_key in customer_cache:
                    customer = customer_cache[cache_key]
                    duplicate_customer_rows_in_file += 1
                else:
                    addr = (address_raw or "").strip() or None
                    customer = Customer(tenant_id=tenant_id, full_name=customer_name, phone=phone, address=addr)
                    session.add(customer)
                    session.flush()
                    customer_cache[cache_key] = customer
                    created_customer_id = customer.id

                job_seq += 1
                ticket_stem = None
                if original_job_id:
                    ticket = original_job_id.strip()
                    if re.match(r"^[\d.]+$", ticket) and ticket.endswith(".0"):
                        ticket = ticket[:-2]
                    if ticket:
                        ticket_stem = f"M{ticket}"
                preferred_num = _allocate_import_job_number(job_number_usage, ticket_stem, job_seq)
                job_number = _reserve_import_job_number(preferred_num, used_job_numbers)

                title = (title_raw.strip() if title_raw else "") or (
                    (work_line[:120] + ("…" if len(work_line) > 120 else "")) if work_line else "Mobile service"
                )
                date_in = _parse_date_flexible(date_in_raw)
                created_at = (
                    datetime(date_in.year, date_in.month, date_in.day, tzinfo=timezone.utc)
                    if date_in
                    else datetime.now(timezone.utc)
                )
                ak_status = _infer_auto_key_status(status_raw, notes_raw)
                quote_cents = _dollars_to_cents(quote_raw)
                cost_cents = _dollars_to_cents(cost_raw)
                job_cost = cost_cents or quote_cents

                vy: int | None = None
                if vehicle_year_raw and str(vehicle_year_raw).strip().isdigit():
                    y = int(str(vehicle_year_raw).strip())
                    if 1950 <= y <= 2100:
                        vy = y

                ca_id = _infer_customer_account_id(session, tenant_id, customer.id)

                enforce_plan_limit(auth, "auto_key_job", auto_key_running_count)
                auto_key_running_count += 1

                job = AutoKeyJob(
                    tenant_id=tenant_id,
                    customer_id=customer.id,
                    customer_account_id=ca_id,
                    job_number=job_number,
                    title=title[:500],
                    description=(notes_raw or None),
                    vehicle_make=vehicle_make or None,
                    vehicle_model=vehicle_model or None,
                    vehicle_year=vy,
                    registration_plate=rego or None,
                    job_address=(address_raw or "").strip() or None,
                    job_type=job_type or None,
                    tech_notes=notes_raw or None,
                    status=ak_status,
                    salesperson=team_member or None,
                    deposit_cents=0,
                    cost_cents=job_cost,
                    created_at=created_at,
                )
                session.add(job)
                session.flush()

                if quote_cents > 0:
                    q_status = "approved" if ak_status in {"completed", "collected", "awaiting_collection"} else "sent"
                    quote = AutoKeyQuote(
                        tenant_id=tenant_id,
                        auto_key_job_id=job.id,
                        status=q_status,
                        subtotal_cents=quote_cents,
                        tax_cents=0,
                        total_cents=quote_cents,
                        currency=tenant_currency,
                        created_at=created_at,
                    )
                    session.add(quote)
                    session.flush()
                    session.add(
                        AutoKeyQuoteLineItem(
                            tenant_id=tenant_id,
                            auto_key_quote_id=quote.id,
                            description="Imported quote total",
                            quantity=1,
                            unit_price_cents=quote_cents,
                            total_price_cents=quote_cents,
                        )
                    )

                if import_log is not None:
                    session.add(
                        ImportLogDetail(
                            import_log_id=import_log.id,
                            row_number=idx,
                            created_customer_id=created_customer_id,
                        )
                    )
                imported += 1

        if dry_run:
            session.rollback()
            return ImportSummaryResponse(
                import_id=uuid4(),
                imported=imported,
                skipped=skipped,
                customers_created=len(customer_cache),
                total_rows=len(rows),
                skipped_reasons=skipped_reasons,
                dry_run=True,
                duplicate_customer_rows_in_file=duplicate_customer_rows_in_file,
                source_sheet=source_sheet,
                import_target=import_target,
            )

        import_log.imported_count = imported
        import_log.skipped_count = skipped
        import_log.customers_created_count = len(customer_cache)
        import_log.status = "completed"
        session.add(import_log)
        session.commit()
    except HTTPException:
        session.rollback()
        raise
    except Exception as exc:
        session.rollback()
        # Full detail goes to logs; clients only see a generic message so we
        # do not leak DB/driver errors, file paths, or stack traces.
        logger.exception("CSV import failed for tenant %s", auth.tenant_id)
        if import_log is not None:
            import_log.status = "failed"
            import_log.error_message = str(exc)
            import_log.imported_count = imported
            import_log.skipped_count = skipped
            import_log.customers_created_count = len(customer_cache)
            session.add(import_log)
            session.commit()
        raise HTTPException(
            status_code=400,
            detail="Import failed. Please check the file and try again.",
        ) from exc

    return ImportSummaryResponse(
        import_id=import_log.id,
        imported=imported,
        skipped=skipped,
        customers_created=len(customer_cache),
        total_rows=len(rows),
        skipped_reasons=skipped_reasons,
        dry_run=False,
        duplicate_customer_rows_in_file=duplicate_customer_rows_in_file,
        source_sheet=source_sheet,
        import_target=import_target,
    )
