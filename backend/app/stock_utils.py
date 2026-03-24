import csv
import io
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl
import xlrd


GROUP_CODE_MAP: dict[str, str] = {
    "AA": "LADIES HEELS",
    "AB": "MENS HEELS",
    "BA": "LADIES SOLES",
    "BC": "MENS SOLES",
    "CA": "MISC BAG REPAIRS",
    "CB": "SNEAKER CLEANING",
    "CD": "MISC SHOE REPAIRS",
    "DA": "FLAT KEYS",
    "DB": "RFID",
    "DC": "AUTO KEY REMOTE",
    "DD": "FUN KEY",
    "DE": "GARAGE GATE REMOTE",
    "EA": "ENGRAVING",
    "EB": "TAGS & PLATES",
    "EC": "COMPUTER ENGRAVING",
    "FB": "SHARPENING",
    "FC": "PHONE SERVICES",
    "FD": "CASH DIFFERENCES",
    "FF": "MISC SERVICES",
    "GA": "SHOE CARE",
    "HA": "KEY RINGS",
    "IA": "LOCKS & SECURITY",
    "JA": "GIFTWARE",
    "JB": "SUNDRY MERCHANDISE",
    "KA": "PROMOTIONAL",
    "KB": "CHARITY",
    "NA": "WATCH REPAIRS",
    "NB": "WATCH BATTERY FITTED",
    "NC": "WATCH BANDS",
    "ND": "3RD PARTY WATCH",
    "NE": "BATTERY SUPPLY ONLY",
    "NF": "OTHER BATTERY FITTED",
}

MASTER_SHEET_PREFERENCE = ("DATA", "SUMMARY")
STOCK_SHEET_PREFERENCE = ("SUMMARY", "DATA", "OPENING")

ITEM_CODE_FIELDS = ("item_code", "itemcode", "code", "stock_code", "sku", "product_code")
GROUP_FIELDS = ("group", "group_code", "category", "category_code")
ITEM_DESCRIPTION_FIELDS = ("item_description", "line_desc", "description", "product_description", "item_desc")
DESCRIPTION_2_FIELDS = ("description_2", "description2", "desc_2", "desc2")
DESCRIPTION_3_FIELDS = ("description_3", "description3", "desc_3", "desc3")
UNIT_DESCRIPTION_FIELDS = ("unit_description", "unit_desc", "unit", "per")
PACK_DESCRIPTION_FIELDS = ("pack_description", "pack_desc", "pack")
PACK_QTY_FIELDS = ("pack_qty", "pack_quantity", "unit_qty_per_pack", "packqty", "unit_qty")
COST_FIELDS = ("price", "item_price", "cost_price", "unit_price", "ord_ex_tax", "order_unit_ex_tax")
RETAIL_FIELDS = ("retail_price", "retail", "sell_price", "rrp", "ord_inc_tax", "ship_inc_tax")
STOCK_FIELDS = ("stock", "stock_qty", "qty_on_hand", "on_hand", "current_stock", "system_stock", "quantity")


def normalize_header(value: Any) -> str:
    text = to_text(value).strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def group_name_for_code(group_code: str | None) -> str | None:
    if not group_code:
        return None
    normalized = group_code.strip().upper()
    return GROUP_CODE_MAP.get(normalized)


def build_full_description(*parts: str | None) -> str:
    seen: set[str] = set()
    ordered: list[str] = []
    for part in parts:
        cleaned = re.sub(r"\s+", " ", (part or "").strip())
        if not cleaned:
            continue
        token = cleaned.casefold()
        if token in seen:
            continue
        seen.add(token)
        ordered.append(cleaned)
    return " | ".join(ordered)


def _first_present(row: dict[str, Any], keys: tuple[str, ...]) -> str:
    for key in keys:
        value = row.get(key)
        if value is not None and to_text(value):
            return to_text(value)
    return ""


def _parse_decimal(value: Any) -> Decimal | None:
    text = to_text(value)
    if not text:
        return None
    cleaned = text.replace(",", "").replace("$", "").strip()
    if cleaned.endswith("%"):
        cleaned = cleaned[:-1].strip()
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = f"-{cleaned[1:-1]}"
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def quantity_from_value(value: Any) -> float | None:
    parsed = _parse_decimal(value)
    if parsed is None:
        return None
    return float(parsed)


def cents_from_value(value: Any) -> int | None:
    parsed = _parse_decimal(value)
    if parsed is None:
        return None
    return int((parsed * Decimal("100")).quantize(Decimal("1")))


def _normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    unnamed_index = 0
    for key, value in row.items():
        header = normalize_header(key)
        if not header:
            unnamed_index += 1
            header = f"unnamed_{unnamed_index}"
        normalized[header] = value
    return normalized


def _csv_rows(raw_bytes: bytes) -> list[dict[str, Any]]:
    text = ""
    for encoding in ("utf-8-sig", "utf-16", "latin-1"):
        try:
            text = raw_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if not text:
        raise ValueError("Unable to decode stock file. Export it as UTF-8 CSV or Excel.")

    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    except csv.Error:
        reader = csv.DictReader(io.StringIO(text))
    return [_normalize_row(row) for row in reader if row]


def _xlsx_sheets(raw_bytes: bytes) -> dict[str, list[dict[str, Any]]]:
    workbook = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
    output: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            continue
        headers = [normalize_header(cell) or f"unnamed_{idx + 1}" for idx, cell in enumerate(values[0])]
        rows: list[dict[str, Any]] = []
        for row in values[1:]:
            record = {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}
            if any(to_text(value) for value in record.values()):
                rows.append(record)
        output[sheet_name.upper()] = rows
    return output


def _xls_sheets(raw_bytes: bytes) -> dict[str, list[dict[str, Any]]]:
    workbook = xlrd.open_workbook(file_contents=raw_bytes)
    output: dict[str, list[dict[str, Any]]] = {}
    for sheet in workbook.sheets():
        if sheet.nrows == 0:
            continue
        headers = [normalize_header(sheet.cell_value(0, col)) or f"unnamed_{col + 1}" for col in range(sheet.ncols)]
        rows: list[dict[str, Any]] = []
        for row_index in range(1, sheet.nrows):
            record = {headers[col]: sheet.cell_value(row_index, col) for col in range(sheet.ncols)}
            if any(to_text(value) for value in record.values()):
                rows.append(record)
        output[sheet.name.upper()] = rows
    return output


def load_stock_sheets(filename: str, raw_bytes: bytes) -> dict[str, list[dict[str, Any]]]:
    extension = Path(filename).suffix.lower()
    if extension == ".csv":
        return {"DATA": _csv_rows(raw_bytes)}
    if extension in {".xlsx", ".xlsm"}:
        return _xlsx_sheets(raw_bytes)
    if extension == ".xls":
        return _xls_sheets(raw_bytes)
    if raw_bytes.startswith(b"PK"):
        return _xlsx_sheets(raw_bytes)
    if raw_bytes.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"):
        return _xls_sheets(raw_bytes)
    return {"DATA": _csv_rows(raw_bytes)}


def _find_group_code(row: dict[str, Any]) -> str:
    explicit = _first_present(row, GROUP_FIELDS).strip().upper()
    if explicit in GROUP_CODE_MAP:
        return explicit
    for key, value in row.items():
        if key.startswith("unnamed_"):
            candidate = to_text(value).strip().upper()
            if candidate in GROUP_CODE_MAP:
                return candidate
    for value in row.values():
        candidate = to_text(value).strip().upper()
        if candidate in GROUP_CODE_MAP:
            return candidate
    return ""


def extract_stock_item_candidate(row: dict[str, Any], *, preferred_sheet: str) -> dict[str, Any] | None:
    item_code = _first_present(row, ITEM_CODE_FIELDS).strip()
    if not item_code:
        return None

    group_code = _find_group_code(row)
    item_description = _first_present(row, ITEM_DESCRIPTION_FIELDS)
    description2 = _first_present(row, DESCRIPTION_2_FIELDS)
    description3 = _first_present(row, DESCRIPTION_3_FIELDS)
    unit_description = _first_present(row, UNIT_DESCRIPTION_FIELDS)
    pack_description = _first_present(row, PACK_DESCRIPTION_FIELDS)
    pack_qty = quantity_from_value(_first_present(row, PACK_QTY_FIELDS))
    cost_price_cents = cents_from_value(_first_present(row, COST_FIELDS))
    retail_price_cents = cents_from_value(_first_present(row, RETAIL_FIELDS))
    system_stock_qty = quantity_from_value(_first_present(row, STOCK_FIELDS))

    return {
        "item_code": item_code.strip(),
        "group_code": group_code,
        "group_name": group_name_for_code(group_code),
        "item_description": item_description or None,
        "description2": description2 or None,
        "description3": description3 or None,
        "full_description": build_full_description(item_description, description2, description3),
        "unit_description": unit_description or None,
        "pack_description": pack_description or None,
        "pack_qty": pack_qty,
        "cost_price_cents": cost_price_cents,
        "retail_price_cents": retail_price_cents,
        "system_stock_qty": system_stock_qty,
        "source_sheet": preferred_sheet,
    }


def merge_stock_records(sheets: dict[str, list[dict[str, Any]]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    items: dict[str, dict[str, Any]] = {}
    source_counts: dict[str, int] = {}

    ordered_sheets: list[str] = []
    for sheet_name in MASTER_SHEET_PREFERENCE + STOCK_SHEET_PREFERENCE:
        if sheet_name not in ordered_sheets and sheet_name in sheets:
            ordered_sheets.append(sheet_name)
    for sheet_name in sheets:
        if sheet_name not in ordered_sheets:
            ordered_sheets.append(sheet_name)

    for sheet_name in ordered_sheets:
        rows = sheets.get(sheet_name, [])
        if not rows:
            continue
        for row in rows:
            candidate = extract_stock_item_candidate(row, preferred_sheet=sheet_name)
            if not candidate:
                continue
            item_code = candidate["item_code"].upper()
            source_counts[sheet_name] = source_counts.get(sheet_name, 0) + 1
            existing = items.get(item_code)
            if not existing:
                items[item_code] = candidate
                continue

            for key in (
                "group_code",
                "group_name",
                "item_description",
                "description2",
                "description3",
                "full_description",
                "unit_description",
                "pack_description",
                "pack_qty",
                "cost_price_cents",
                "retail_price_cents",
            ):
                if candidate.get(key) not in (None, "", 0):
                    existing[key] = candidate[key]

            if candidate.get("system_stock_qty") is not None:
                existing["system_stock_qty"] = candidate["system_stock_qty"]

            if not existing.get("source_sheet") or existing.get("source_sheet") == "SUMMARY":
                existing["source_sheet"] = candidate["source_sheet"]

    ordered_items = sorted(items.values(), key=lambda item: (item.get("group_code") or "ZZ", item["item_code"]))
    meta = {
        "sheet_names": list(sheets.keys()),
        "source_counts": source_counts,
        "imported_items": len(ordered_items),
    }
    return ordered_items, meta
