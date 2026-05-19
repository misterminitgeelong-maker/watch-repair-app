"""Parse Mister Minit shop rows from TSS / shop-list Excel exports."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any, BinaryIO, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# Default local path for the Dec 2025 TSS export (do not commit the workbook).
DEFAULT_TSS_XLSX_PATH = r"c:\Users\samme\Downloads\TSS Dec25 Report (1).xlsx"

_SHOP_HEADER_MARKERS = ("shop #", "shop#")
_SUBHEADER_MARKERS = ("raw score", "tss score")
_PREFERRED_SHEET_NAMES = ("shops", "tss scores")

# TSS column mapping (sheet "TSS Scores", header row with "Shop #"):
# | Excel column | Field        | Example        |
# |--------------|--------------|----------------|
# | Shop #       | shop_number  | 3269           |
# | Shop Name    | name         | Chadstone      |
# | Area         | area         | VIC SOUTH      |
# | Region       | region       | VIC, SW, NZ, … |
#
# Region values in the Dec 2025 export:
#   VIC, NSW, QLD — Australian state codes (matches shops in that state)
#   SW            — South-West cluster (WA + South Australia areas)
#   NZ            — New Zealand (areas NZ NORTH / NZ SOUTH)
#   SEA           — South-East Asia (areas e.g. MALAYSIA)
#
# AU state for filtering is derived from Area when Region is a cluster code (SW/SEA/NZ).

_AU_STATE_AREA_PREFIXES: tuple[tuple[str, str], ...] = (
    ("SOUTH AUSTRALIA", "SA"),
    ("VIC", "VIC"),
    ("NSW", "NSW"),
    ("QLD", "QLD"),
    ("WA", "WA"),
    ("TAS", "TAS"),
    ("NT", "NT"),
    ("ACT", "ACT"),
)
_AU_STATE_REGION_CODES = frozenset({"VIC", "NSW", "QLD", "WA", "SA", "TAS", "NT", "ACT"})


@dataclass
class MinitShopsParseResult:
    shops: list[MinitShopRow]
    errors: list[str] = field(default_factory=list)
    sheet_name: str | None = None


@dataclass(frozen=True)
class MinitShopRow:
    shop_number: str
    name: str
    area: str | None
    region: str | None  # TSS Region column (VIC, SW, NZ, SEA, …)

    @property
    def business_address(self) -> str:
        parts = [p for p in (self.name, self.area, self.region) if p]
        return ", ".join(parts)

    @property
    def state_code(self) -> str | None:
        """Best AU state code for routing/filtering (derived from Area, then Region)."""
        return derive_au_state_from_area_region(self.area, self.region)


def derive_au_state_from_area_region(area: str | None, region: str | None) -> str | None:
    if area:
        upper = area.strip().upper()
        for prefix, code in _AU_STATE_AREA_PREFIXES:
            if upper == prefix or upper.startswith(f"{prefix} "):
                return code
    if region:
        code = region.strip().upper()
        if code in _AU_STATE_REGION_CODES:
            return code
    return None


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_shop_number(raw: Any) -> str | None:
    text = _cell_str(raw)
    if not text or text.lower() in {"nan", "shop #"}:
        return None
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    if not re.fullmatch(r"\d{1,10}", text):
        return None
    return text


def _find_header_row(rows: Iterable[tuple[Any, ...]]) -> int | None:
    for idx, row in enumerate(rows):
        first = _cell_str(row[0] if row else "").lower()
        if first in _SHOP_HEADER_MARKERS:
            return idx
    return None


def _row_is_subheader(row: tuple[Any, ...]) -> bool:
    joined = " ".join(_cell_str(c).lower() for c in row[:6])
    return any(m in joined for m in _SUBHEADER_MARKERS)


def _sheet_has_shop_header(ws: Worksheet, *, scan_rows: int = 30) -> bool:
    for row in ws.iter_rows(max_row=scan_rows, values_only=True):
        first = _cell_str(row[0] if row else "").lower()
        if first in _SHOP_HEADER_MARKERS:
            return True
    return False


def _resolve_worksheet(wb: Workbook, sheet_name: str | None) -> Worksheet:
    if sheet_name:
        return wb[sheet_name]
    lowered = {name.strip().lower(): name for name in wb.sheetnames}
    for preferred in _PREFERRED_SHEET_NAMES:
        if preferred in lowered:
            return wb[lowered[preferred]]
    for name in wb.sheetnames:
        ws = wb[name]
        if _sheet_has_shop_header(ws):
            return ws
    return wb.active


def _parse_rows_from_worksheet(
    ws: Worksheet,
    *,
    source_label: str,
    collect_row_errors: bool,
) -> MinitShopsParseResult:
    raw_rows = list(ws.iter_rows(values_only=True))
    header_idx = _find_header_row(raw_rows)
    if header_idx is None:
        raise ValueError(f"No 'Shop #' header row found in {source_label}")

    headers = [_cell_str(c) for c in raw_rows[header_idx]]
    col_index = {h.lower(): i for i, h in enumerate(headers) if h}

    def col(name: str, fallback: int) -> int:
        return col_index.get(name.lower(), fallback)

    i_shop = col("shop #", col("shop#", 0))
    i_name = col("shop name", 1)
    i_area = col("area", 2)
    i_region = col("region", 3)

    shops: list[MinitShopRow] = []
    errors: list[str] = []
    seen: set[str] = set()

    for row_num, row in enumerate(raw_rows[header_idx + 1 :], start=header_idx + 2):
        if not row or _row_is_subheader(row):
            continue
        raw_shop = row[i_shop] if i_shop < len(row) else None
        shop_number = _normalize_shop_number(raw_shop)
        name = _cell_str(row[i_name] if i_name < len(row) else "")
        raw_shop_text = _cell_str(raw_shop)

        if raw_shop_text and not shop_number:
            if collect_row_errors:
                errors.append(f"Row {row_num}: invalid shop number '{raw_shop_text}'")
            continue
        if shop_number and not name:
            if collect_row_errors:
                errors.append(f"Row {row_num}: shop {shop_number} missing name")
            continue
        if not shop_number or not name:
            continue
        if shop_number in seen:
            if collect_row_errors:
                errors.append(f"Row {row_num}: duplicate shop number {shop_number} (skipped)")
            continue
        seen.add(shop_number)
        area = _cell_str(row[i_area] if i_area < len(row) else "") or None
        region = _cell_str(row[i_region] if i_region < len(row) else "") or None
        shops.append(
            MinitShopRow(
                shop_number=shop_number,
                name=name,
                area=area,
                region=region,
            )
        )

    return MinitShopsParseResult(shops=shops, errors=errors, sheet_name=ws.title)


def parse_minit_shops_xlsx(
    path: str | Path | None = None,
    *,
    file_obj: BinaryIO | bytes | None = None,
    sheet_name: str | None = None,
    collect_row_errors: bool = False,
) -> list[MinitShopRow]:
    """Load shop rows from a Minit shop-list workbook (header row contains ``Shop #``)."""
    return parse_minit_shops_xlsx_detailed(
        path,
        file_obj=file_obj,
        sheet_name=sheet_name,
        collect_row_errors=collect_row_errors,
    ).shops


def parse_minit_shops_xlsx_detailed(
    path: str | Path | None = None,
    *,
    file_obj: BinaryIO | bytes | None = None,
    sheet_name: str | None = None,
    collect_row_errors: bool = False,
) -> MinitShopsParseResult:
    """Parse workbook and return shops plus optional per-row error messages."""
    if path is None and file_obj is None:
        raise ValueError("path or file_obj is required")
    source = Path(path) if path is not None else "uploaded workbook"
    load_target: str | Path | BytesIO
    if file_obj is not None:
        data = file_obj.read() if hasattr(file_obj, "read") else file_obj
        load_target = BytesIO(data)
    else:
        load_target = Path(path)  # type: ignore[arg-type]

    wb = load_workbook(load_target, read_only=True, data_only=True)
    try:
        ws = _resolve_worksheet(wb, sheet_name)
        return _parse_rows_from_worksheet(
            ws,
            source_label=str(source),
            collect_row_errors=collect_row_errors,
        )
    finally:
        wb.close()


def tenant_slug_for_shop(shop: MinitShopRow, *, prefix: str = "minit") -> str:
    """Stable slug: minit-{shop_number} (unique, DNS-safe)."""
    return f"{prefix}-{shop.shop_number}"
