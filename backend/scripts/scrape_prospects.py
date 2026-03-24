#!/usr/bin/env python3
"""
Raw prospect scraper: fetches businesses from directory sites and exports to Excel.
No Google API — uses direct HTTP requests and HTML parsing.

Usage:
  cd backend && python -m scripts.scrape_prospects --state VIC --limit 10 -o prospects.xlsx
  python -m scripts.scrape_prospects --category mechanics --state NSW -o mechanics_nsw.xlsx

Options:
  --state     Limit to one state (e.g. VIC, NSW)
  --category  Limit to one category
  --limit     Max suburbs to process (0 = no limit)
  --delay     Seconds between requests (default: 2)
  -o, --output  Output Excel file path (default: prospects_<timestamp>.xlsx)
  --dry-run   Show what would be scraped, don't fetch or write
"""
from __future__ import annotations

import argparse
import re
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlencode

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import httpx
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlmodel import Session, select

from app.database import engine
from app.models import Suburb
from app.routes.prospects import CATEGORY_BASES, STATE_CODES, _state_name_for_query

DEFAULT_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml",
    "Accept-Language": "en-AU,en;q=0.9",
}

# Directory sources
YP_SEARCH = "https://www.yellowpages.com.au/search/listings"
TL_BASE = "https://www.truelocal.com.au"


def build_yp_url(clue: str, location: str, page: int = 1) -> str:
    """Build Yellow Pages search URL."""
    params: dict = {"clue": clue, "locationClue": location}
    if page > 1:
        params["page"] = str(page)
    return f"{YP_SEARCH}?{urlencode(params)}"


def parse_yp_listings(html: str) -> list[dict]:
    """Extract business listings from YP search results HTML."""
    soup = BeautifulSoup(html, "html.parser")
    results: list[dict] = []

    # YP often uses listing cards with various structures; try common patterns
    # Class names may change — adjust if site structure updates
    cards = soup.select('[data-listing-id], .listing-card, .listing, [class*="listing"]')
    if not cards:
        cards = soup.select('article, .result, [class*="result"]')
    if not cards:
        # Fallback: look for links to /business/
        cards = soup.select('a[href*="/business/"]')

    seen: set[str] = set()
    for card in cards[:50]:  # Limit per page
        name = ""
        address = ""
        phone = ""
        website = ""

        # Try to get business name
        name_el = card.select_one(
            "h2 a, h3 a, .listing-name a, [class*='name'] a, a[href*='/business/']"
        )
        if name_el:
            name = (name_el.get_text(strip=True) or "").strip()
        if not name and card.name == "a":
            name = (card.get_text(strip=True) or "").strip()

        # Address
        addr_el = card.select_one(
            "[class*='address'], [class*='location'], address, .address"
        )
        if addr_el:
            address = (addr_el.get_text(strip=True) or "").strip()

        # Phone
        phone_el = card.select_one('a[href^="tel:"]')
        if phone_el:
            phone = (phone_el.get("href", "").replace("tel:", "").strip() or "").strip()
        else:
            phone_match = re.search(r"\b\d{2}\s?\d{4}\s?\d{4}\b", card.get_text() or "")
            if phone_match:
                phone = phone_match.group(0)

        # Website
        site_el = card.select_one('a[href*="http"]:not([href^="tel:"])')
        if site_el and "yellowpages" not in (site_el.get("href") or "").lower():
            website = (site_el.get("href") or "").strip()

        if name and name not in seen:
            seen.add(name)
            results.append({"name": name, "address": address, "phone": phone, "website": website})

    return results


def build_truelocal_url(category_slug: str, location_slug: str) -> str:
    """Build True Local find URL: /find/{category}/{location}."""
    return f"{TL_BASE}/find/{category_slug}/{location_slug}"


# Map our categories to True Local category slugs
CATEGORY_TO_TL_SLUG = {
    "car_dealerships": "car-dealers",
    "used_car_dealers": "used-car-dealers",
    "car_rental": "car-rental",
    "mechanics": "mechanics",
    "panel_beaters": "panel-beaters",
    "insurance": "car-insurance",
    "fleet_management": "fleet-management",
    "car_auctions": "car-auctions",
}


def parse_truelocal_listings(html: str) -> list[dict]:
    """Extract business listings from True Local HTML."""
    soup = BeautifulSoup(html, "html.parser")
    results: list[dict] = []
    cards = soup.select("article, .listing, [class*='listing'], [class*='business']")
    if not cards:
        cards = soup.select('a[href*="/business/"]')
    seen: set[str] = set()
    for card in cards[:50]:
        name = ""
        address = ""
        phone = ""
        website = ""
        name_el = card.select_one("h2 a, h3 a, .name a, a[href*='/business/']")
        if name_el:
            name = (name_el.get_text(strip=True) or "").strip()
        if not name and card.name == "a":
            name = (card.get_text(strip=True) or "").strip()
        addr_el = card.select_one("[class*='address'], [class*='location']")
        if addr_el:
            address = (addr_el.get_text(strip=True) or "").strip()
        phone_el = card.select_one('a[href^="tel:"]')
        if phone_el:
            phone = (phone_el.get("href", "").replace("tel:", "").strip() or "").strip()
        else:
            m = re.search(r"\b\d{2}\s?\d{4}\s?\d{4}\b", card.get_text() or "")
            if m:
                phone = m.group(0)
        site_el = card.select_one('a[href*="http"]:not([href^="tel:"])')
        if site_el and "truelocal" not in (site_el.get("href") or "").lower():
            website = (site_el.get("href") or "").strip()
        if name and name not in seen:
            seen.add(name)
            results.append({"name": name, "address": address, "phone": phone, "website": website})
    return results


def scrape_yp(
    client: httpx.Client | None,
    clue: str,
    location: str,
    *,
    max_pages: int = 2,
    use_browser: bool = False,
) -> list[dict]:
    """Scrape Yellow Pages for a search term and location."""
    all_results: list[dict] = []
    seen_names: set[str] = set()

    for page in range(1, max_pages + 1):
        url = build_yp_url(clue, location, page)
        html: str | None = None
        if use_browser:
            html = fetch_with_browser(url)
        elif client:
            try:
                resp = client.get(url, headers=DEFAULT_HEADERS, timeout=15.0)
                resp.raise_for_status()
                html = resp.text
            except Exception as e:
                print(f"    Warning: {e}")
                break
        if not html:
            break
        listings = parse_yp_listings(html)
        for L in listings:
            if L["name"] not in seen_names:
                seen_names.add(L["name"])
                all_results.append(L)

        if len(listings) < 10:  # No more pages
            break

    return all_results


def scrape_truelocal(
    client: httpx.Client | None,
    category_slug: str,
    location_slug: str,
    *,
    use_browser: bool = False,
) -> list[dict]:
    """Scrape True Local for a category and location."""
    url = build_truelocal_url(category_slug, location_slug)
    html: str | None = None
    if use_browser:
        html = fetch_with_browser(url)
    elif client:
        try:
            resp = client.get(url, headers=DEFAULT_HEADERS, timeout=15.0)
            resp.raise_for_status()
            html = resp.text
        except Exception as e:
            print(f"    Warning: {e}")
            return []
    if not html:
        return []
    return parse_truelocal_listings(html)


def slugify(s: str) -> str:
    """Convert to URL slug: lowercase, spaces to hyphens."""
    return re.sub(r"[^a-z0-9-]+", "-", s.lower().strip()).strip("-")


def fetch_with_browser(url: str) -> str | None:
    """Fetch HTML using Playwright (bypasses many bot blocks). Requires: pip install playwright && playwright install chromium"""
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            page.set_extra_http_headers(DEFAULT_HEADERS)
            page.goto(url, wait_until="domcontentloaded", timeout=15000)
            html = page.content()
            browser.close()
            return html
    except ImportError:
        print("  Install Playwright for --browser mode: pip install playwright && playwright install chromium")
        return None
    except Exception as e:
        print(f"  Browser fetch failed: {e}")
        return None


def main() -> int:
    parser = argparse.ArgumentParser(description="Scrape prospects from directories to Excel")
    parser.add_argument("--state", type=str, required=True, help="State code (e.g. VIC, NSW)")
    parser.add_argument("--category", type=str, help="Limit to one category")
    parser.add_argument("--limit", type=int, default=0, help="Max suburbs (0 = no limit)")
    parser.add_argument("--delay", type=float, default=2.0, help="Seconds between requests")
    parser.add_argument("-o", "--output", type=str, help="Output Excel path")
    parser.add_argument(
        "--source",
        choices=["yp", "truelocal"],
        default="truelocal",
        help="Directory to scrape",
    )
    parser.add_argument(
        "--browser",
        action="store_true",
        help="Use Playwright (pip install playwright && playwright install chromium). Bypasses 403 on many sites.",
    )
    parser.add_argument("--dry-run", action="store_true", help="Don't fetch or write")
    args = parser.parse_args()

    state = args.state.strip().upper()
    if state not in STATE_CODES:
        print(f"Invalid state: {args.state}")
        return 1

    categories = [args.category] if args.category else list(CATEGORY_BASES.keys())
    if args.category and args.category not in CATEGORY_BASES:
        print(f"Invalid category. Choose from: {', '.join(CATEGORY_BASES.keys())}")
        return 1

    # Get suburbs from DB or fallback list
    suburb_names: list[str] = []
    try:
        with Session(engine) as session:
            q = select(Suburb).where(Suburb.state_code == state).order_by(Suburb.name)
            rows = list(session.exec(q).all())
            suburb_names = [r.name for r in rows]
    except Exception:
        pass
    if not suburb_names:
        from app.routes.prospects import SUBURBS_BY_STATE_FALLBACK
        suburb_names = SUBURBS_BY_STATE_FALLBACK.get(state, [])
    suburbs = [type("Sub", (), {"name": n, "state_code": state})() for n in suburb_names]

    if args.limit:
        suburbs = suburbs[: args.limit]

    state_name = _state_name_for_query(state)
    total_queries = len(suburbs) * len(categories)
    print(f"Scraping: {len(categories)} category(ies) × {len(suburbs)} suburbs = {total_queries} URLs")
    print(f"Output: {args.output or 'prospects_<timestamp>.xlsx'}")

    if args.dry_run:
        print("DRY RUN — no requests or file writes")
        return 0

    all_prospects: list[dict] = []
    done = 0

    use_browser = args.browser
    client = httpx.Client(timeout=20.0, follow_redirects=True) if not use_browser else None
    try:
        for suburb in suburbs:
            sub_name = suburb.name
            location = f"{sub_name} {state_name}"
            for cat in categories:
                try:
                    if args.source == "yp":
                        clue = CATEGORY_BASES[cat]
                        rows = scrape_yp(client, clue, location, use_browser=use_browser)
                    else:
                        cat_slug = CATEGORY_TO_TL_SLUG.get(cat, slugify(CATEGORY_BASES[cat].replace(" ", "-")))
                        loc_slug = slugify(f"{sub_name} {state}")  # e.g. melbourne-vic
                        rows = scrape_truelocal(client, cat_slug, loc_slug, use_browser=use_browser)
                    for r in rows:
                        r["category"] = cat
                        r["suburb"] = sub_name
                        r["state"] = state
                        all_prospects.append(r)
                except Exception as e:
                    print(f"  Error {cat} / {sub_name}: {e}")
                done += 1
                if done % 5 == 0:
                    print(f"  Progress: {done}/{total_queries} | found: {len(all_prospects)}")
                time.sleep(args.delay)
    finally:
        if client:
            client.close()

    out_path = args.output
    if not out_path:
        stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
        out_path = f"prospects_{state}_{stamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Prospects"
    headers = ["Name", "Address", "Phone", "Website", "Category", "Suburb", "State"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, p in enumerate(all_prospects, 2):
        ws.cell(row=row_idx, column=1, value=p.get("name", ""))
        ws.cell(row=row_idx, column=2, value=p.get("address", ""))
        ws.cell(row=row_idx, column=3, value=p.get("phone", ""))
        ws.cell(row=row_idx, column=4, value=p.get("website", ""))
        ws.cell(row=row_idx, column=5, value=p.get("category", ""))
        ws.cell(row=row_idx, column=6, value=p.get("suburb", ""))
        ws.cell(row=row_idx, column=7, value=p.get("state", ""))
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    wb.save(out_path)
    print(f"Saved {len(all_prospects)} prospects to {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
